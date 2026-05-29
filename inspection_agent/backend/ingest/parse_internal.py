"""把 data/internal/ 下的合同/报告/SOP 全部解析成纯文本,写到 data/internal_text.jsonl。

支持格式:
  .doc / .docx  → 用 Word COM 抽(用户机器装了 Word)
  .xlsx         → openpyxl
  .xls          → xlrd
  .pdf          → 单独处理(报告 PDF 走 OCR,见 probe_report_pdf.py 后续脚本)
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import win32com.client

ROOT = Path(__file__).resolve().parent.parent.parent
INTERNAL_DIR = ROOT / "data" / "internal"
OUTPUT = ROOT / "data" / "internal_text.jsonl"

CATEGORIES = ["contracts", "reports", "sops"]


def parse_word(path: Path, word_app) -> str:
    """Word COM 抽 .doc / .docx 的全文。"""
    doc = word_app.Documents.Open(str(path.resolve()), ReadOnly=True)
    try:
        return doc.Content.Text or ""
    finally:
        doc.Close(SaveChanges=False)


def parse_xlsx(path: Path) -> str:
    """.xlsx 抽全部 sheet → 制表符分隔的文本。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join("" if c is None else str(c) for c in row)
            if row_text.strip():
                parts.append(row_text)
    wb.close()
    return "\n".join(parts)


def parse_xls(path: Path) -> str:
    """.xls 旧格式,用 xlrd。"""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    parts = []
    for sheet in wb.sheets():
        parts.append(f"=== Sheet: {sheet.name} ===")
        for r in range(sheet.nrows):
            row = sheet.row_values(r)
            row_text = "\t".join(str(c) for c in row)
            if row_text.strip():
                parts.append(row_text)
    return "\n".join(parts)


def collect_files():
    files = []
    for cat in CATEGORIES:
        d = INTERNAL_DIR / cat
        if not d.exists():
            continue
        for f in sorted(d.iterdir()):
            if not f.is_file():
                continue
            if f.name.startswith("."):
                continue
            files.append((cat, f))
    return files


def main():
    files = collect_files()
    if not files:
        print(f"[!] {INTERNAL_DIR} 下没有可处理的文件")
        sys.exit(1)

    print(f"[*] 共 {len(files)} 个内部材料文件")
    OUTPUT.parent.mkdir(exist_ok=True, parents=True)

    word_app = None
    n_ok, n_skip, n_fail = 0, 0, 0

    out = OUTPUT.open("w", encoding="utf-8")
    try:
        for cat, f in files:
            ext = f.suffix.lower()
            tag = f"  [{cat}] {f.name}"
            print(tag, end=" ... ", flush=True)
            try:
                if ext in (".doc", ".docx"):
                    if word_app is None:
                        print("\n    [*] 启动 Word COM...")
                        word_app = win32com.client.Dispatch("Word.Application")
                        word_app.Visible = False
                        word_app.DisplayAlerts = False
                        print(tag, end=" ... ", flush=True)
                    text = parse_word(f, word_app)
                elif ext == ".xlsx":
                    text = parse_xlsx(f)
                elif ext == ".xls":
                    text = parse_xls(f)
                elif ext == ".pdf":
                    print("跳过(PDF 单独处理)")
                    n_skip += 1
                    continue
                else:
                    print(f"跳过(未支持 {ext})")
                    n_skip += 1
                    continue

                record = {
                    "category": cat,
                    "source_file": f.name,
                    "ext": ext,
                    "char_count": len(text),
                    "text": text,
                }
                out.write(json.dumps(record, ensure_ascii=False) + "\n")
                print(f"OK ({len(text)} 字)")
                n_ok += 1
            except Exception as e:
                print(f"[X] {e}")
                n_fail += 1
    finally:
        out.close()
        if word_app:
            try:
                word_app.Quit()
            except Exception:
                pass

    print(f"\n[OK] 写入 {OUTPUT}")
    print(f"     成功 {n_ok} / 跳过 {n_skip} / 失败 {n_fail}")


if __name__ == "__main__":
    main()
