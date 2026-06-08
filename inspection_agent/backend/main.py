"""
检测行业 Agent — 工业级 FastAPI 后端
SSE 流式 API + API Key 认证 + 限流 + 日志
"""
from __future__ import annotations

import json
import os
import smtplib
import sys
import time
from collections import defaultdict
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Iterable

from fastapi import FastAPI, HTTPException, UploadFile, File, Request, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from pydantic import BaseModel

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import (
    save_conversation, list_conversations, delete_conversation,
    log_usage, get_usage_stats, init_db,
)

os.environ.setdefault("HF_HUB_OFFLINE", "1")
os.environ.setdefault("TRANSFORMERS_OFFLINE", "1")

from agent import build_agent  # noqa: E402

# ========== 配置 ==========
ROOT = Path(__file__).resolve().parent.parent
FEEDBACK_FILE = ROOT / "data" / "feedback.jsonl"

# API Key 认证
API_KEY = os.environ.get("AGENT_API_KEY", "")
# CORS 允许的域名（生产环境必须限制）
CORS_ORIGINS = os.environ.get("AGENT_CORS_ORIGINS", "*").split(",")
# 文件上传限制
MAX_FILE_SIZE_MB = int(os.environ.get("AGENT_MAX_FILE_MB", "10"))
MAX_FILE_SIZE = MAX_FILE_SIZE_MB * 1024 * 1024
# 限流配置
RATE_LIMIT_WINDOW = 60          # 窗口秒数
RATE_LIMIT_MAX = 30             # 每窗口最大请求数

app = FastAPI(title="检测行业 Agent")

# CORS — 生产环境必须限制域名
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in CORS_ORIGINS if o.strip()],
    allow_methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# ========== 请求日志中间件 ==========
@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    client = request.client.host if request.client else "-"
    response = await call_next(request)
    elapsed = (time.time() - start) * 1000
    print(f"[{datetime.now().isoformat()}] {client} {request.method} {request.url.path} {response.status_code} {elapsed:.1f}ms")
    return response


# ========== 限流（内存版，单实例够用） ==========
_rate_limit_store = defaultdict(list)  # ip -> [timestamp, ...]


def check_rate_limit(ip: str) -> bool:
    now = time.time()
    window_start = now - RATE_LIMIT_WINDOW
    # 清理过期记录
    _rate_limit_store[ip] = [t for t in _rate_limit_store[ip] if t > window_start]
    if len(_rate_limit_store[ip]) >= RATE_LIMIT_MAX:
        return False
    _rate_limit_store[ip].append(now)
    return True


# ========== API Key 认证 ==========
def verify_api_key(x_api_key: str = Header(None, alias="X-API-Key")):
    """如果配置了 API_KEY，则必须提供正确的 Key"""
    if not API_KEY:
        return True  # 未配置时不校验（开发环境）
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="无效的 API Key")
    return True


# ========== Agent 启动 ==========
agent = None


@app.on_event("startup")
def startup():
    global agent
    print("[*] 启动时构建 Agent ...")
    agent = build_agent()
    print("[OK] Agent 就绪")
    print(f"[*] API_KEY {'已配置' if API_KEY else '未配置（开发模式）'}")
    print(f"[*] CORS: {CORS_ORIGINS}")
    print(f"[*] 文件上限: {MAX_FILE_SIZE_MB}MB")


# ========== Pydantic 模型 ==========
class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]


class ConversationSave(BaseModel):
    id: str
    title: str
    messages: list[dict]


class FeedbackRequest(BaseModel):
    message: str
    rating: int | None = None
    name: str | None = None
    related_query: str | None = None


# ========== 工具函数 ==========
def _serialize_tool_call(tc: dict[str, Any]) -> dict:
    return {"type": "tool_call", "name": tc.get("name", ""), "args": tc.get("args", {})}


def _serialize_tool_result(content: str) -> dict:
    if isinstance(content, str) and len(content) > 1500:
        content = content[:1500] + "..."
    return {"type": "tool_result", "content": content}


def _sse(data: dict) -> str:
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


def stream_agent(messages_payload: list[tuple[str, str]]) -> Iterable[str]:
    yield _sse({"type": "start"})
    try:
        for chunk in agent.stream({"messages": messages_payload}, stream_mode="updates"):
            for node, state in chunk.items():
                msgs = state.get("messages", []) if isinstance(state, dict) else []
                if not msgs:
                    continue
                msg = msgs[-1]
                tool_calls = getattr(msg, "tool_calls", None)
                if tool_calls:
                    for tc in tool_calls:
                        yield _sse(_serialize_tool_call(tc))
                    continue
                if node == "tools":
                    yield _sse(_serialize_tool_result(getattr(msg, "content", "")))
                    continue
                content = getattr(msg, "content", "")
                if content:
                    yield _sse({"type": "text", "content": content})
    except Exception as e:
        yield _sse({"type": "error", "message": str(e)})
    yield _sse({"type": "done"})


# ========== 路由 ==========
@app.get("/health")
def health():
    return {"ok": True, "agent_ready": agent is not None, "time": datetime.now().isoformat()}


@app.post("/chat")
def chat(req: ChatRequest, request: Request, _=Depends(verify_api_key)):
    # 限流检查
    client_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown")
    if not check_rate_limit(client_ip):
        raise HTTPException(status_code=429, detail="请求过于频繁，请稍后再试")

    if agent is None:
        raise HTTPException(503, "Agent 未就绪")
    if not req.messages:
        raise HTTPException(400, "messages 不能为空")
    if req.messages[-1].role != "user":
        raise HTTPException(400, "最后一条 message 必须是 user")

    payload = [(m.role, m.content) for m in req.messages]
    question = req.messages[-1].content[:300]

    def logged_stream():
        tools_called = []
        for chunk in stream_agent(payload):
            if '"type":"tool_call"' in chunk:
                import re
                m = re.search(r'"name":"([^"]+)"', chunk)
                if m:
                    tools_called.append(m.group(1))
            yield chunk
        try:
            log_usage(question, tools_called, 0, client_ip)
        except Exception:
            pass

    return StreamingResponse(
        logged_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ========== 文件上传 ==========
def is_image(ext: str) -> bool:
    return ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp")


def extract_text(file_path: str, filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".txt", ".md", ".json", ".csv", ".py", ".js", ".ts", ".tsx", ".jsx", ".html", ".css", ".xml", ".yaml", ".yml", ".log"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(file_path)
            text = "".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            return "[PDF 解析库未安装]"
    if ext in (".docx", ".doc"):
        try:
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return "[Word 解析库未安装]"
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        text += row_text + "\n"
                if len(text) > 50000:
                    text = text[:50000] + "\n... (文件过长,已截断)"
            wb.close()
            return text
        except Exception as e:
            return f"[Excel 解析失败: {e}]"
    if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
        return "[图片文件，请使用 analyze_image 工具分析]"
    return f"[不支持的文件类型: {ext}]"


@app.post("/upload")
async def upload_file(file: UploadFile = File(...), _=Depends(verify_api_key)):
    if not file.filename:
        raise HTTPException(400, "文件名不能为空")

    # 大小检查
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, f"文件过大，最大允许 {MAX_FILE_SIZE_MB}MB")

    ext = os.path.splitext(file.filename)[1].lower()
    allowed = {".txt", ".md", ".pdf", ".docx", ".doc", ".xlsx", ".xls",
               ".json", ".csv", ".py", ".js", ".ts", ".tsx", ".html",
               ".css", ".xml", ".yaml", ".yml", ".log", ".png", ".jpg",
               ".jpeg", ".gif", ".bmp", ".webp"}
    if ext not in allowed:
        raise HTTPException(400, f"不支持的文件类型: {ext}")

    if is_image(ext):
        upload_dir = ROOT / "data" / "uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        img_path = upload_dir / safe_name
        img_path.write_bytes(content)
        image_url = f"http://1.15.170.85/inspection/uploads/{safe_name}"
        return {"filename": file.filename, "ext": ext, "text": f"[图片文件] 链接: {image_url}", "image_url": image_url, "length": 0}

    import tempfile
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        text = extract_text(tmp_path, file.filename)
    finally:
        os.unlink(tmp_path)

    return {"filename": file.filename, "ext": ext, "text": text, "length": len(text)}


# ========== 静态文件服务 ==========
@app.get("/uploads/{filename}")
def serve_upload(filename: str):
    file_path = ROOT / "data" / "uploads" / filename
    if file_path.exists():
        return FileResponse(file_path)
    raise HTTPException(404, "图片不存在")


@app.get("/contracts/{filename}")
def serve_contract(filename: str):
    file_path = ROOT / "data" / "contracts" / filename
    if file_path.exists():
        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename=filename)
    raise HTTPException(404, "文件不存在")


# ========== 对话记忆 ==========
@app.post("/conversations")
def save_conv(req: ConversationSave, _=Depends(verify_api_key)):
    try:
        save_conversation(req.id, req.title, req.messages)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/conversations")
def list_convs(_=Depends(verify_api_key)):
    return list_conversations()


@app.delete("/conversations/{conv_id}")
def delete_conv(conv_id: str, _=Depends(verify_api_key)):
    delete_conversation(conv_id)
    return {"ok": True}


# ========== 用量统计 ==========
@app.get("/admin/stats")
def admin_stats(_=Depends(verify_api_key)):
    return get_usage_stats()


# ========== Feedback ==========
def _send_feedback_email(entry: dict) -> None:
    host = os.getenv("SMTP_HOST")
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASSWORD")
    if not (host and user and password):
        raise RuntimeError("SMTP 未配置")
    port = int(os.getenv("SMTP_PORT", "465"))
    to = os.getenv("SMTP_TO", user)
    rating = entry.get("rating")
    rating_str = f"{rating}★" if rating else "无评分"
    name = entry.get("name") or "匿名"
    body_lines = [
        f"时间: {entry['timestamp']}",
        f"评分: {rating_str}",
        f"反馈人: {name}",
        f"相关查询: {entry.get('related_query') or '无'}",
        "",
        "===== 反馈内容 =====",
        entry["message"],
    ]
    body = "\n".join(body_lines)
    msg = MIMEMultipart()
    msg["From"] = user
    msg["To"] = to
    msg["Subject"] = f"[检测Agent反馈] {name} · {rating_str}"
    msg.attach(MIMEText(body, "plain", "utf-8"))
    with smtplib.SMTP_SSL(host, port, timeout=10) as server:
        server.login(user, password)
        server.send_message(msg)


@app.post("/feedback")
def feedback(req: FeedbackRequest):
    msg = (req.message or "").strip()
    if not msg:
        raise HTTPException(400, "反馈内容不能为空")
    if len(msg) > 2000:
        raise HTTPException(400, "反馈内容过长(最多 2000 字)")

    entry = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "message": msg,
        "rating": req.rating,
        "name": (req.name or "").strip() or None,
        "related_query": (req.related_query or "").strip() or None,
    }
    FEEDBACK_FILE.parent.mkdir(exist_ok=True, parents=True)
    with FEEDBACK_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    email_status = "skipped"
    email_error = None
    try:
        _send_feedback_email(entry)
        email_status = "sent"
    except RuntimeError:
        pass
    except Exception as e:
        email_status = "failed"
        email_error = str(e)
        print(f"[WARN] 邮件发送失败: {e}")

    return {"ok": True, "saved": True, "email": email_status, "email_error": email_error}


# 前端静态文件
FRONTEND_DIR = ROOT / "dist"

@app.get("/")
def root():
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    return JSONResponse({
        "name": "检测行业 Agent API",
        "version": "2.0.0-industrial",
        "endpoints": {
            "POST /chat": "Chat with the inspection agent (SSE streaming, requires X-API-Key header)",
            "POST /upload": "Upload file for analysis (requires X-API-Key header)",
            "POST /feedback": "Submit user feedback",
            "GET /health": "Health check",
        },
    })


@app.get("/{filename}")
def static_frontend(filename: str):
    """Serve frontend static files"""
    file_path = FRONTEND_DIR / filename
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)
    # SPA fallback: return index.html for unknown routes
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    raise HTTPException(404, "文件不存在")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
