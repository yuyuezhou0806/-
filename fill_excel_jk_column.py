#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
填写合同清单报价表J列和K列
J列 = J列所在/覆盖的行对应的I列数值之和（如果是合并单元格，计算整个范围；如果是独立单元格，就是I列值）
K列 = J列 × F列
"""

import openpyxl
from openpyxl import load_workbook
import os


def is_number(value):
    """检查是否为有效数字"""
    if value is None or value == '-' or value == '':
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def get_cell_value(sheet_data, row, col):
    """获取单元格的计算值（使用data_only模式加载的工作簿）"""
    try:
        cell = sheet_data.cell(row=row, column=col)
        return cell.value
    except:
        return None


def get_j_cell_info(sheet, row):
    """获取J列单元格的信息（是否是合并单元格，范围是什么）"""
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_col <= 10 <= merged_range.max_col:
            if merged_range.min_row <= row <= merged_range.max_row:
                # 这个行在J列的合并单元格范围内
                return {
                    'is_merged': True,
                    'is_main_cell': row == merged_range.min_row,
                    'main_row': merged_range.min_row,
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row
                }
    # 不在任何合并单元格范围内
    return {
        'is_merged': False,
        'is_main_cell': True,
        'main_row': row,
        'start_row': row,
        'end_row': row
    }


def process_excel(filepath):
    """处理Excel文件，填写J列和K列"""

    if not os.path.exists(filepath):
        print(f"错误：文件不存在 {filepath}")
        return False

    print(f"正在加载文件: {filepath}")

    # 用data_only模式读取计算后的值
    wb_data = load_workbook(filepath, data_only=True)
    sheet_data = wb_data.active

    # 用普通模式打开用于写入
    wb = load_workbook(filepath)
    sheet = wb.active

    print(f"工作表: {sheet.title}, 总行数: {sheet.max_row}")

    # 存储J列计算结果（用于K列计算）
    j_values = {}

    # ========== 第一步：填写J列 ==========
    print("\n正在填写J列...")

    # 先收集所有J列合并单元格的信息，避免重复计算
    processed_merged_ranges = set()

    for row in range(5, sheet.max_row + 1):
        j_info = get_j_cell_info(sheet, row)

        if j_info['is_merged']:
            range_key = (j_info['start_row'], j_info['end_row'])

            if range_key in processed_merged_ranges:
                # 这个合并单元格已经处理过了，跳过
                continue

            processed_merged_ranges.add(range_key)

            # 计算整个合并范围内I列的和
            i_sum = 0
            i_details = []
            for r in range(j_info['start_row'], j_info['end_row'] + 1):
                i_value = get_cell_value(sheet_data, r, 9)  # I列是第9列
                if is_number(i_value):
                    i_num = float(i_value)
                    i_sum += i_num
                    i_details.append(f"I{r}={i_num}")

            # 填入主单元格
            main_cell = sheet.cell(row=j_info['main_row'], column=10)  # J列是第10列
            main_cell.value = i_sum

            # 记录该范围内所有行的J值（用于K列计算）
            for r in range(j_info['start_row'], j_info['end_row'] + 1):
                j_values[r] = i_sum

            print(f"  J{j_info['start_row']}-J{j_info['end_row']} (合并): {', '.join(i_details[:5])}{'...' if len(i_details) > 5 else ''} -> 总和={i_sum}")

        else:
            # 独立单元格，直接取I列值
            i_value = get_cell_value(sheet_data, row, 9)
            if is_number(i_value):
                j_val = float(i_value)
            else:
                j_val = 0

            cell = sheet.cell(row=row, column=10)
            cell.value = j_val
            j_values[row] = j_val

            if j_val != 0:
                print(f"  J{row} (独立): I{row}={j_val}")

    # ========== 第二步：填写K列 ==========
    print("\n正在填写K列 (K = J × F)...")

    k_filled_count = 0
    for row in range(5, sheet.max_row + 1):
        j_val = j_values.get(row, 0)
        f_value = get_cell_value(sheet_data, row, 6)  # F列是第6列

        if is_number(f_value):
            f_val = float(f_value)
        else:
            f_val = 0

        # 计算 K = J × F
        if j_val != 0 and f_val != 0:
            k_val = j_val * f_val

            # 填入K列（第11列）
            k_cell = sheet.cell(row=row, column=11)
            k_cell.value = k_val
            k_filled_count += 1

            if k_filled_count <= 20:  # 只打印前20个
                print(f"  K{row}: J{row}({j_val}) × F{row}({f_val}) = {k_val}")

    print(f"  共填写 {k_filled_count} 个K列单元格")

    # 保存文件（加时间戳避免冲突）
    from datetime import datetime
    timestamp = datetime.now().strftime('%H%M%S')
    output_path = filepath.replace('.xlsx', f'_已填写_{timestamp}.xlsx')
    wb.save(output_path)
    print(f"\n完成！已保存到: {output_path}")

    # 关闭workbook
    wb_data.close()
    wb.close()

    return True


if __name__ == '__main__':
    # 文件路径
    filepath = r"C:\Users\admin\Desktop\附件4：合同清单报价-标段一-zch111.xlsx"

    process_excel(filepath)
