import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

list_all={
    '宋建华' : {'张菊':'辅助工','章琦仕':'检测','马绍斌':'检测','胡富强':'检测','顾书华':'检测','何欢':'检测','吴耀龙':'检测',
                '戴斌':'检测','陈婷凤':'检测','柯晨':'检测','杨浦':'辅助工','张颖':'辅助工','李元明':'检测',
                '李朕':'检测','史安平':'检测','徐勤雨':'检测','蔡韦迪':'检测','唐云巍':'检测','卢海华':'检测','钱春林':'辅助工',
                '陈永彬':'辅助工','花兴伯':'辅助工','金培浩':'检测','李树叶':'检测','刘贵山':'检测','钟钱顺':'检测',
                '徐双虎':'检测','胡涛':'检测','胡吉祥':'检测','吕静':'核算'},
    '朱晓辉' : {'张菊' : '辅助工','章琦仕':'检测','马绍斌':'检测','胡富强':'检测','顾书华':'检测','何欢':'检测','吴耀龙':'检测',
                '戴斌':'检测','陈婷凤':'检测','柯晨':'检测','杨浦':'辅助工','张颖':'辅助工','李元明':'检测',
                '李朕':'检测','史安平':'检测','徐勤雨':'检测','蔡韦迪':'检测','唐云巍':'检测','卢海华':'检测','钱春林':'辅助工',
                '陈永彬':'辅助工','花兴伯':'辅助工','金培浩':'检测','李树叶':'检测','刘贵山':'检测','钟钱顺':'检测',
                '徐双虎':'检测','胡涛':'检测','胡吉祥':'检测','吕静':'核算'},
    '陈凌溦' : {'何永忠':'辅助工','吴登红':'辅助工','赵春海':'辅助工','王倩男':'检测','胡杰新':'检测','谢世轩':'检测','张闻昭':'辅助工','丛远才':'检测','尤丹':'检测','李朕':'检测','梅剑韬':'检测','刘芳':'检测','林明洁':'检测','吴建明':'检测','李元彬':'辅助工','胡木军':'辅助工','陈侃':'辅助工','杜汉文':'检测','袁荣见':'检测','朱长兴':'检测','金新月':'检测','张超群':'检测','李磊':'检测','黎丽艳':'检测','吕静':'核算'},
    '李富刚' : {'何永忠':'辅助工','吴登红':'辅助工','赵春海':'辅助工','王倩男':'检测','胡杰新':'检测','谢世轩':'检测','张闻昭':'辅助工','丛远才':'检测','尤丹':'检测','李朕':'检测','梅剑韬':'检测','刘芳':'检测','林明洁':'检测','吴建明':'检测','李元彬':'辅助工','胡木军':'辅助工','陈侃':'辅助工','杜汉文':'检测','袁荣见':'检测','朱长兴':'检测','金新月':'检测','张超群':'检测','李磊':'检测','黎丽艳':'检测','吕静':'核算'},
    '朱渊士' : {'谯爱民':'检测','陈强':'检测','王倩男':'检测','周洋':'检测','龙斌':'检测','杨杨':'检测','刁雨浩':'检测','张佰林':'辅助工','季俊芬':'辅助工','倪志强':'辅助工','查超':'辅助工','李潇潇':'检测','黄宇峰':'检测','胡海强':'检测','高杰':'检测','徐强':'检测','胡杰':'检测','李孝平':'检测','胡涛':'检测','秦昌凯':'检测','张德乾':'检测','李曼':'辅助工','徐加军':'辅助工','徐浩淋':'辅助工','朱晔':'辅助工','刘晨阳':'辅助工','康凯':'辅助工','吕静':'核算'},
    '马峰' : {'谯爱民':'检测','陈强':'检测','王倩男':'检测','周洋':'检测','龙斌':'检测','杨杨':'检测','刁雨浩':'检测','张佰林':'辅助工','季俊芬':'辅助工','倪志强':'辅助工','查超':'辅助工','李潇潇':'检测','黄宇峰':'检测','胡海强':'检测','高杰':'检测','徐强':'检测','胡杰':'检测','李孝平':'检测','胡涛':'检测','秦昌凯':'检测','张德乾':'检测','李曼':'辅助工','徐加军':'辅助工','徐浩淋':'辅助工','朱晔':'辅助工','刘晨阳':'辅助工','康凯':'辅助工','吕静':'核算'},
    '丁家杨' : {'秦命国':'辅助工','潘佳华':'检测','金岩':'检测','杨俊':'检测','朱清生':'检测','郭伟丽':'检测','沈彤':'检测','益伟红':'检测','秦桐':'辅助工','刘志':'辅助工','于佳明':'检测','高思凯':'检测','周蓉清':'检测','杨铭杰':'检测','刘海林':'检测','汪先明':'检测','孙福志':'辅助工','于永华':'辅助工','汪鹏':'检测','莫子犇':'检测','袁爱林':'检测','吴昊':'辅助工','刘海斌':'检测','施佳杰':'检测','程金龙':'检测','庞紫玄':'检测','肖华山':'检测','吕静':'核算'},
    '裘增辉' : {'秦命国':'辅助工','潘佳华':'检测','金岩':'检测','杨俊':'检测','朱清生':'检测','郭伟丽':'检测','沈彤':'检测','益伟红':'检测','秦桐':'辅助工','刘志':'辅助工','于佳明':'检测','高思凯':'检测','周蓉清':'检测','杨铭杰':'检测','刘海林':'检测','汪先明':'检测','孙福志':'辅助工','于永华':'辅助工','汪鹏':'检测','莫子犇':'检测','袁爱林':'检测','吴昊':'辅助工','刘海斌':'检测','施佳杰':'检测','程金龙':'检测','庞紫玄':'检测','肖华山':'检测','吕静':'核算'},
    '沈文荣' : {'潘建伟':'辅助工','高锦华':'辅助工','郑阔昌':'检测','李皛':'检测','谢林灵':'检测','周平':'检测','戴斌':'检测','李贺':'检测','宋鑫涛':'检测','王学宇':'辅助工','贾富':'辅助工','黄伟':'辅助工','朱亦飞':'辅助工','张猛':'检测','袁建波':'检测','董月灯':'检测','李家宝':'检测','宋俊杰':'检测','常彦荣':'检测','周晓华':'检测','朱建明':'辅助工','朱天玉':'辅助工','李先友':'检测','徐一丰':'检测','徐绪绪':'检测','王泽男':'辅助工','袁鹏清':'检测','陈康强':'检测','吕静':'核算'},
    '石蕾' : {'潘建伟':'辅助工','高锦华':'辅助工','郑阔昌':'检测','李皛':'检测','谢林灵':'检测','周平':'检测','戴斌':'检测','李贺':'检测','宋鑫涛':'检测','王学宇':'辅助工','贾富':'辅助工','黄伟':'辅助工','朱亦飞':'辅助工','张猛':'检测','袁建波':'检测','董月灯':'检测','李家宝':'检测','宋俊杰':'检测','常彦荣':'检测','周晓华':'检测','朱建明':'辅助工','朱天玉':'辅助工','李先友':'检测','徐一丰':'检测','徐绪绪':'检测','王泽男':'辅助工','袁鹏清':'检测','陈康强':'检测','吕静':'核算'},
    '王少峰' : {'潘建伟':'辅助工','高锦华':'辅助工','郑阔昌':'检测','李皛':'检测','谢林灵':'检测','周平':'检测','戴斌':'检测','李贺':'检测','宋鑫涛':'检测','王学宇':'辅助工','贾富':'辅助工','黄伟':'辅助工','朱亦飞':'辅助工','张猛':'检测','袁建波':'检测','董月灯':'检测','李家宝':'检测','宋俊杰':'检测','常彦荣':'检测','周晓华':'检测','朱建明':'辅助工','朱天玉':'辅助工','李先友':'检测','徐一丰':'检测','徐绪绪':'检测','王泽男':'辅助工','袁鹏清':'检测','陈康强':'检测','吕静':'核算'},
    '卫天勇' : {'李纪奎':'辅助工','洪登国':'辅助工','钱凡':'检测','茅佳成':'检测','刘大伟':'检测','蒲建江':'辅助工','丁文斌':'辅助工','何迟应':'辅助工','祝志永':'检测','李英':'检测','朱冬进':'检测','杜晓蔚':'检测','李朕':'检测','谢冬进':'检测','胡杰':'检测','陈江邺':'检测','戚明星':'检测','王国文':'辅助工','朱志伟':'检测','徐善建':'检测','徐双虎':'检测','黄鑫':'检测','刘宇晨':'检测','杨昊':'检测','胡涛':'检测','李真':'检测','张萍':'检测','朱庆一':'辅助工','胡祥勇':'检测','吕静':'核算'},
    '沈健' : {'李纪奎':'辅助工','洪登国':'辅助工','钱凡':'检测','茅佳成':'检测','刘大伟':'检测','蒲建江':'辅助工','丁文斌':'辅助工','何迟应':'辅助工','祝志永':'检测','李英':'检测','朱冬进':'检测','杜晓蔚':'检测','李朕':'检测','谢冬进':'检测','胡杰':'检测','陈江邺':'检测','戚明星':'检测','王国文':'辅助工','朱志伟':'检测','徐善建':'检测','徐双虎':'检测','黄鑫':'检测','刘宇晨':'检测','杨昊':'检测','胡涛':'检测','李真':'检测','张萍':'检测','朱庆一':'辅助工','胡祥勇':'检测','吕静':'核算'},
    '凡前龙' : {'王泳太':'辅助工','陆威':'检测','陈晖':'检测','韩明喜':'检测','郝思荣':'检测','江小国':'检测','宋辉':'检测','熊建宇':'检测','刘兵辉':'辅助工','许前聪':'检测','宋德强':'检测','罗杰':'检测','黄宇峰':'检测','李静志':'检测','刘彦梅':'检测','刘大朋':'检测','姚秋名':'辅助工','王宇':'检测','蔡建军':'检测','方贤华':'检测','李高飞':'检测','娄庆华':'检测','仇慧':'检测','孙晓磊':'检测','吕静':'核算'},
    '张勇' : {'王泳太':'辅助工','陆威':'检测','陈晖':'检测','韩明喜':'检测','郝思荣':'检测','江小国':'检测','宋辉':'检测','熊建宇':'检测','刘兵辉':'辅助工','许前聪':'检测','宋德强':'检测','罗杰':'检测','黄宇峰':'检测','李静志':'检测','刘彦梅':'检测','刘大朋':'检测','姚秋名':'辅助工','王宇':'检测','蔡建军':'检测','方贤华':'检测','李高飞':'检测','娄庆华':'检测','仇慧':'检测','金':'检测'},
    '秦昌凯' : {'刘志':'辅助工','郭朱霞':'检测','庞有发':'辅助工','何欢':'检测','王凯远':'检测','戴晓峰':'检测','王志翔':'检测','王永琪':'辅助工','吕静':'核算'},
    '朱萍' : {'刘志':'辅助工','庞有发':'辅助工','何欢':'检测','王凯远':'检测','戴晓峰':'检测','王志翔':'检测','王永琪':'辅助工','吕静':'核算'},
    '尤丹' : {'刘志':'辅助工','庞有发':'辅助工','何欢':'检测','王凯远':'检测','戴晓峰':'检测','王志翔':'检测','王永琪':'辅助工','吕静':'核算'},
    '娄庆华' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '曹江' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '戴斌' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '朱志伟' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '谢冬进' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '钱凡' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '江小国' : {'樊斌':'辅助工','沈庭威':'辅助工','仇建明':'检测','陈广春':'检测','孙亚建':'检测','吕静':'核算'},
    '卢海华' : {'张豹':'检测','王静宇':'检测','吕静':'核算'},
    '徐双虎' : {'张豹':'检测','王静宇':'检测','吕静':'核算'},
    '刘彦梅' : {'张豹':'检测','王静宇':'检测','吕静':'核算'},
    '李真' : {'张辉':'辅助工','殷福俊':'辅助工','孟佳壕':'辅助工','张怡俊':'辅助工','杨江江':'检测','刘浩':'辅助工','刘亚':'检测','聂礼鹏':'检测','谷鹏轩':'检测','郭飞洋':'检测','吕静':'核算'},
    '常彦荣' : {'张辉':'辅助工','殷福俊':'辅助工','孟佳壕':'辅助工','张怡俊':'辅助工','杨江江':'检测','刘浩':'辅助工','刘亚':'检测','聂礼鹏':'检测','谷鹏轩':'检测','郭飞洋':'检测','吕静':'核算'},
    '刘大伟' : {'张辉':'辅助工','殷福俊':'辅助工','孟佳壕':'辅助工','张怡俊':'辅助工','杨江江':'检测','刘浩':'辅助工','刘亚':'检测','聂礼鹏':'检测','谷鹏轩':'检测','郭飞洋':'检测','吕静':'核算'},
    '莫子犇' : {'张辉':'辅助工','殷福俊':'辅助工','孟佳壕':'辅助工','张怡俊':'辅助工','杨江江':'检测','刘浩':'辅助工','刘亚':'检测','聂礼鹏':'检测','谷鹏轩':'检测','郭飞洋':'检测','吕静':'核算'},
    '王少峰，凡前龙，环境，结构，零星，' : {'潘建伟':'辅助工','高锦华':'辅助工','郑阔昌':'检测','李皛':'检测','谢林灵':'检测','周平':'检测','戴斌':'检测','李贺':'检测','宋鑫涛':'检测','王学宇':'辅助工','贾富':'辅助工','黄伟':'辅助工','朱亦飞':'辅助工','张猛':'检测','袁建波':'检测','董月灯':'检测','李家宝':'检测','宋俊杰':'检测','常彦荣':'检测','周晓华':'检测','朱建明':'辅助工','朱天玉':'辅助工','李先友':'检测','徐一丰':'检测','徐绪绪':'检测','王泽男':'辅助工','袁鹏清':'检测','陈康强':'检测','吕静':'核算','刘志':'辅助工','郭朱霞':'检测','庞有发':'辅助工','何欢':'检测','王凯远':'检测','戴晓峰':'检测','王志翔':'检测','王永琪':'辅助工','王泳太':'辅助工','陆威':'检测','陈晖':'检测','韩明喜':'检测','郝思荣':'检测','江小国':'检测','宋辉':'检测','熊建宇':'检测','刘兵辉':'辅助工','许前聪':'检测','宋德强':'检测','罗杰':'检测','黄宇峰':'检测','李静志':'检测','刘彦梅':'检测','刘大朋':'检测','姚秋名':'辅助工','王宇':'检测','蔡建军':'检测','方贤华':'检测','李高飞':'检测','娄庆华':'检测','仇慧':'检测','孙晓磊':'检测',"沈庭威": "辅助工","丁文斌": "辅助工","花兴伯": "辅助工","张德乾": "检测","徐善建": "检测","胡杰新": "检测","林明洁": "检测","汪鹏": "检测"},}
def add_worker(driver, no, name, job):
    no = no + 2
    # 1 点击白框
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='childTableBody-ProjectTeamEntry']/tr[{}]/td[3]/div/div".format(no)).click()
    # 2 点击铅笔
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='childTableBody-ProjectTeamEntry']/tr[{}]/td[2]/nz-input-group/span/span[2]".format(no)).click()
    # 3 点击对应角色
    time.sleep(1)
    if job == '辅助工':
        driver.find_element(By.XPATH, "//*[@id='body']/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[2]/td[2]").click()
    elif job == '检测':
        driver.find_element(By.XPATH, "//*[@id='body']/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[4]/td[2]").click()
    elif job == '核算':
        driver.find_element(By.XPATH,
                            "//*[@id='body']/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[5]/td[2]").click()
    else:
        raise Exception('职位数据异常！检查全局变量和数据文件！')
    # 4 点击确定
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]").click()
    # 5 点击空白
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='childTableBody-ProjectTeamEntry']/tr[{}]/td[2]/div/div".format(no)).click()
    # 6 点击铅笔二
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='childTableBody-ProjectTeamEntry']/tr[{}]/td[3]/nz-input-group/span/span[2]".format(no)).click()
    # 7 输入名称
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='body']/nz-input-group/input").send_keys(name)
    time.sleep(1)
    driver.find_element(By.XPATH, "//*[@id='body']/nz-input-group/input").send_keys(Keys.ENTER)
    # 8 点击对应人员
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[2]").click()
    # 9 确定
    time.sleep(1)
    driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]").click()


def main_point(project_num, project_owner):
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://192.168.99.91/')
        driver.maximize_window()
        driver.find_element(By.CLASS_NAME, 'ant-input').send_keys('Z0343')
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button').click()
        # 3 自建应用
        time.sleep(10)
        el0 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span')))
        el0.click()
        # 4 项目管理
        time.sleep(3)
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-1"]/div/div/div/nz-spin/div/div[1]/div[4]').click()
        # 5 项目立项
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[4]/a').click()
        # 6 项目预算单
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[4]/ul/li[1]/a').click()
        # 7 查询方案
        time.sleep(5)
        el1 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[1]/div/div/button[2]')))
        el1.click()
        # 8 加号图标
        time.sleep(2)
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[1]/div/div/a[1]').click()
        # driver.find_element(By.CLASS_NAME,'ant-btn-default').click()
        # 9 【过滤条件：项目编码，比较符：等于，比较值：项目编码】
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/query-set/pop-window[1]/div/div/div[2]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[3]/nz-select/div/div').click()
        time.sleep(3)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/ul/li[10]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[4]/nz-select/div/div').click()
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-4"]/div/div/ul/li[1]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').send_keys(project_num)
        # 10 确定
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/query-set/pop-window[1]/div/div/div[3]/div/div/button[1]').click()
        # 11 小方块
        time.sleep(2)
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[1]/label/span[1]/input").click()
        print(1)
        # 12 关联生成
        time.sleep(2)
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[1]/div/div/button[7]").click()
        # 13 小加号
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/pop-window[1]/div/div/div[2]/div/div/nz-tree/ul/nz-tree-node[2]/li/span[1]").click()
        # 14 成员表
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/pop-window[1]/div/div/div[2]/div/div/nz-tree/ul/nz-tree-node[2]/li/ul/nz-tree-node/li/span[2]/span").click()
        # 15 确定
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/pop-window[1]/div/div/div[3]/button[1]").click()
        time.sleep(2)

        # 16 获取对应leader的worker数据和数量
        try:
            list_worker = list_all[project_owner]
            list_len = len(list_worker)
        except:
            raise Exception('leader数据异常！检查全局变量和数据文件！')

        # 17 点很多下新增数据
        for i in range(list_len):
            driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[8]/div/div/div/div/div/nz-tabset/div[2]/div/div/div/div/div/div/div/button").click()

        # 18 依据全局变量填充数据
        for i in list_worker:
            add_worker(driver, list(list_worker).index(i)+1, i ,list_worker[i])

        # 19 审核
        time.sleep(2)
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[3]/div/nz-form-control/div/span/button[3]").click()
        print('finish')
        time.sleep(5)
        driver.close()
        return True
    except Exception:
        return False

def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')

def yyz_main(main_excel_name):
    log_file_name = 'log_' + main_excel_name + '.txt'
    error_data = []
    wb=openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = wb.active
    project_max = sheet.max_row
    print('共{}条数据'.format(project_max))
    for i in range(project_max):
        project_num = sheet.cell(i+1,1).value
        project_owner = sheet.cell(i+1,2).value
        start_log = 'No:{}, project_num:{}, project_owner"{}, start_time:{}'.format(i + 1, project_num, project_owner, time.strftime('%Y-%m-%d %H:%M:%S'))
        print(start_log)
        write_log_data(log_file_name, start_log)
        flag = main_point(project_num, project_owner)
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')

    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(project_max, project_max - fail_data_sum, fail_data_sum)
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, i)


if __name__ == '__main__':
    excel_name = input('表格名：')
    yyz_main(excel_name)
