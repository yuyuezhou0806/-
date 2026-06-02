from datetime import datetime
import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main_point(project_num, workload):
    """
    :param project_num:项目编号
    :param workload:原因
    :return: 布尔型，返回运行结果
    """
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://bpm.chinajyy.net/#/passport/login')
        driver.maximize_window()
        driver.find_element(By.CLASS_NAME, 'ant-input').send_keys('Z0343')
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button').click()
        # 3，自建应用
        time.sleep(5)
        el0 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span')))
        el0.click()
        # 4，项目管理
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-1"]/div/div/div/nz-spin/div/div[1]/div[4]/small').click()
        # 6，项目开工
        time.sleep(5)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[5]/a').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[5]/ul/li[1]/a').click()
        time.sleep(2)
        # 7 查询
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[1]/div/div/button[2]').click()
        time.sleep(2)
        # 8 选择项目编码
        driver.find_element(By.XPATH,'//*[@id="body"]/nz-layout/nz-sider/div/ul/li[2]').click()
        time.sleep(2)
        # 9 清楚格内容,填入内容
        driver.find_element(By.XPATH,'//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').clear()
        time.sleep(1)
        driver.find_element(By.XPATH,
                            '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').click()
        driver.find_element(By.XPATH,
                            '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').send_keys(project_num)
        time.sleep(1)
        driver.find_element(By.XPATH,'//*[@id="footer"]/div/div[2]/button[1]').click()
        time.sleep(2)
        # 10 勾选内容,关联生成表
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[1]/label/span[1]/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[1]/div/div/button[7]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'//*[@id="body"]/div/div/nz-tree/ul/nz-tree-node/li/span[1]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'//*[@id="body"]/div/div/nz-tree/ul/nz-tree-node/li/ul/nz-tree-node/li/span[2]/span').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/pop-window[1]/div/div/div[3]/button[1]').click()
        # 14，填入变更原因
        time.sleep(2)
        driver.find_element(By.XPATH,'//*[@id="body"]/div[2]/div/form/div[14]/div/div/div/div/div/div/div/nz-form-control/div/span/textarea').send_keys(workload)
        # 16,填入新日期
        time.sleep(2)
        driver.find_element(By.XPATH,
                            '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[16]/div/div/div/div/div/nz-tabset/div[2]/div/div/div/div/div/div/div/nz-table/nz-spin/div/div/div/div/div/table/tbody/tr/td[6]/nz-date-picker/nz-picker/span/input').click()
        time.sleep(2)
        # 选择年份td列tr行
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/inner-popup/calendar-header/div/div/span/a[1]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/inner-popup/calendar-header/div/year-panel/div/div/div[2]/table/tbody/tr[3]/td[2]/a').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/inner-popup/calendar-header/div/div/span/a[2]').click()
        time.sleep(1)
        # 选择月份
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/inner-popup/calendar-header/div/month-panel/div/div/div[2]/month-table/table/tbody/tr[4]/td[3]/a').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/inner-popup/div/date-table/table/tbody/tr[5]/td[4]/div').click()
        time.sleep(1)
        # 17,抓取调整天数
        b = driver.find_element(By.XPATH, '//*[@id="childTableBody-XmgqtzdEntry"]/tr/td[7]/div/div').text
        print(b)
        driver.find_element(By.XPATH,'//*[@id="body"]/div[2]/div/form/div[13]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input').send_keys(b)
        time.sleep(1)
        # 18,延期类型
        driver.find_element(By.XPATH,'//*[@id="childTableBody-XmgqtzdEntry"]/tr/td[8]/div/div').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="childTableBody-XmgqtzdEntry"]/tr/td[8]/nz-select/div/div').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/div/ul/li[2]').click()
        time.sleep(1)
        el = driver.find_element(By.XPATH,
                                 '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[18]/div/div/div/div/div/div/div/nz-form-control/div/span/div[1]/div[1]/div/nz-upload/div/div/input')
        el.send_keys(r"C:\Users\admin\PyCharmMiscProject\1226延期.pdf")
        time.sleep(5)
        # 23，审核
        driver.find_element(By.XPATH, '//*[@id="body"]/div[3]/div[2]/nz-form-control/div/span/button[2]').click()
        print('finish')
        time.sleep(5)
        driver.close()
        return True
    except Exception:
        return False


def read_excel(main_excel_name):
    a = []
    workbook = openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = workbook.active
    for i in sheet.iter_rows():
        for j in i:
            a.append(j.value)
    return a


def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')


def yyz_main(main_excel_name):
    # 把表格名拼接出日志文件名
    log_file_name = 'log_' + main_excel_name + '.txt'
    # 建空列表记录运行错误的编号
    error_data = []
    # 读excel
    wb = openpyxl.load_workbook(main_excel_name + '.xlsx')
    # 默认sheet，如果指定sheet名就 sheet = wb.active('sheet名')
    sheet = wb.active
    # 因为第一行是标题，获取最大长度要-1，比如10行的excel就是9个数据,test_num=9
    test_num = sheet.max_row - 1
    # 从2开始遍历到9+2，因为python是左闭右开，所以代码是[2,11),从第二行读到第十行,一共9次
    for i in range(2, test_num + 2):
        # 获取对应行数和第1列的项目编号，和第2列的工作量
        project_num = sheet.cell(row=i, column=1).value
        workload = sheet.cell(row=i, column=2).value
        # 打日志
        print('共{}条数据'.format(test_num))
        start_log = 'No:{},project_num:{},workload:{},start_time:{}'.format(i - 1, project_num, workload, time.strftime('%Y-%m-%d %H:%M:%S'))
        print(start_log)
        write_log_data(log_file_name, start_log)
        # 运行selenium脚本，把这次运行的project_num和workload传进去，返回类型为布尔型，True或者False
        flag = main_point(project_num, workload)
        # 按返回类型打日志是否运行成功
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            # 把运行错误的编号存到error_data列表
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')
    # 那列表长度确定错误数
    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(test_num,
                                                                                                 test_num - fail_data_sum,
                                                                                                 fail_data_sum)
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, i)


if __name__ == '__main__':
    # 输入获取表格名，传到主函数
    excel_name = input('表格名：')
    yyz_main(excel_name)
