from datetime import datetime
import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main_point(project_num):
    """
    :param project_num:项目编号
    :param workload:需汇报工作量
    :param col3_data:人工工时
    :return: 布尔型，返回运行结果
    """
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://192.168.99.91/')
        driver.find_element(By.CLASS_NAME, 'ant-input').send_keys('Z0343')
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button').click()
        # 3，自建应用
        time.sleep(5)
        el0 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span')))
        el0.click()
        # 4，项目管理
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-1"]/div/div/div/nz-spin/div/div[1]/div[3]').click()
        # 6，查询方案
        time.sleep(5)
        el1 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[1]/div[1]/button[1]')))
        el1.click()
        # 7，加号图标
        time.sleep(3)
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[1]/div/div/a[1]').click()
        # driver.find_element(By.CLASS_NAME,'ant-btn-default').click()
        # 8，【过滤条件：项目编码，比较符：等于，比较值：项目编码】
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[3]/nz-select/div').click()
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-3"]/div/div/ul/li[1]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[4]/nz-select/div/div').click()
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-4"]/div/div/ul/li[1]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').send_keys(project_num)
        # 9，确定按钮
        driver.find_element(By.XPATH, '//*[@id="footer"]/div/div/button[1]').click()
        # 10，选择项目编码对应项目
        time.sleep(6)
        el2 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[2]')))
        el2.click()
        # 11，关联生成
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[1]/div[1]/nz-dropdown[1]/button').click()
        time.sleep(1)
        # 12，生成项目工作汇报单
        driver.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/ul[3]/li/div').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[5]/div/div/ul/ul[1]/li/a').click()
        # 16，新增数据
        time.sleep(3)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[18]/div/div/div/div/div/div/div/nz-form-control/div/span/nz-select/div/div/div[1]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/div/ul/li[3]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,"/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[19]/div/div/div/div/div/nz-tabset/div[2]/div[1]/div/div/div/div/div/div/button").click()
        # 17，点击空白框
        print(3)
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr/td[2]/div/div').click()
        time.sleep(1)
        # 18，点击编辑图标
        driver.find_element(By.XPATH,'//*[@id="childTableBody-ProWorkReportEW"]/tr/td[2]/nz-input-group/span/span[2]').click()
        # 19，选择人
        time.sleep(3)
        worker_element = driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[1]/td[1]')
        ActionChains(driver).double_click(worker_element).perform()
        time.sleep(1)
        # 20，点击合计工时框
        driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr/td[5]/div/div').click()
        # 21，输入工时
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr/td[5]/nz-input-number/div[2]/input').send_keys('0')
        # 22,输入工作量
        time.sleep(1)
        driver.find_element(By.XPATH,"//*[@id='body']/div[2]/div/form/div[11]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input").click()
        driver.find_element(By.XPATH,"//*[@id='body']/div[2]/div/form/div[11]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input").send_keys("0")
        time.sleep(1)
        # 23，审核ZCHJC003869
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[3]/div/nz-form-control/div/span/button[3]').click()
        print('finish')
        time.sleep(5)
        driver.close()
        return True
    except Exception:
        return False


def read_excel(main_excel_name):
    a = []
    workbook = openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = workbook.active
    for i in sheet.iter_rows():
        for j in i:
            a.append(j.value)
    return a


def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')


def yyz_main(main_excel_name):
    # 把表格名拼接出日志文件名
    log_file_name = 'log_' + main_excel_name + '.txt'
    # 建空列表记录运行错误的编号
    error_data = []
    # 读excel
    wb = openpyxl.load_workbook(main_excel_name + '.xlsx')
    # 默认sheet，如果指定sheet名就 sheet = wb.active('sheet名')
    sheet = wb.active
    # 因为第一行是标题，获取最大长度要-1，比如10行的excel就是9个数据,test_num=9
    test_num = sheet.max_row - 1
    # 从2开始遍历到9+2，因为python是左闭右开，所以代码是[2,11),从第二行读到第十行,一共9次
    for i in range(2, test_num + 2):
        # 获取对应行数和第1列的项目编号，和第2列的工作量
        project_num = sheet.cell(row=i, column=1).value
        # 打日志
        print('共{}条数据'.format(test_num))
        start_log = 'No:{},project_num:{},start_time:{}'.format(i - 1, project_num,
                                                                            time.strftime('%Y-%m-%d %H:%M:%S'))
        print(start_log)
        write_log_data(log_file_name, start_log)
        # 运行selenium脚本，把这次运行的project_num和workload传进去，返回类型为布尔型，True或者False
        flag = main_point(project_num)
        # 按返回类型打日志是否运行成功
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            # 把运行错误的编号存到error_data列表
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')
    # 那列表长度确定错误数
    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(test_num,
                                                                                                 test_num - fail_data_sum,
                                                                                                 fail_data_sum)
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, i)


if __name__ == '__main__':
    # 输入获取表格名，传到主函数
    excel_name = input('表格名：')
    yyz_main(excel_name)