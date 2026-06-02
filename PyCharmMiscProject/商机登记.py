import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main_point(project_num, workload, col3_data):
    """
    :param project_num:工程名称
    :param workload:客户名称
    :param col3_data:金额
    :return: 布尔型，返回运行结果
    """
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://192.168.99.91/')
        driver.maximize_window()
        driver.find_element(By.CLASS_NAME, 'ant-input').send_keys('Z0209')
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button').click()
        # 3，自建应用
        time.sleep(5)
        el0 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span')))
        el0.click()
        # 4，CRM,商机管理，商机登记单
        driver.find_element(By.XPATH, '/html/body/div[3]/div[3]/div/div/div/div/nz-spin/div/div[1]/div[3]/i').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[4]/a').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/layout-sidebar/div/sidebar-nav/ul/li[4]/ul/li[4]/a').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/div/as-split/as-split-area[2]/div[1]/div/div/button[3]').click()
        time.sleep(2)
        # 5,点击商机名称，输入商机名称,编辑，输入客户，输入金额
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[2]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[2]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/input').send_keys(project_num)
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[16]/div/div/div/div/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-input-group/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-input-group/input').send_keys(workload)
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-input-group/input').send_keys(Keys.ENTER)
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[1]/td[1]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[8]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[8]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input').send_keys(col3_data)
        # 6，子公司
        time.sleep(2)
        driver.find_element(By.XPATH,'//*[@id="body"]/div[2]/div/form/div[4]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'//*[@id="body"]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[16]/td[2]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        # 7.业务承接所属
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[5]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'//*[@id="body"]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[1]/td[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        # 8.投资额或面积
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[8]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[8]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/input').send_keys('/')
        time.sleep(1)
        # 9.华东区域
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[9]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[3]/td[1]').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        time.sleep(1)
        # 10.定标日期
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[11]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-date-picker/nz-picker/span/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-4"]/div/date-range-popup/div/div/div/div/calendar-footer/div/span/today-button/a').click()
        time.sleep(1)
        # 11.开标日期，中标可能性
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[12]/div/div/div/div/div/div/div/nz-form-control/div/span/nz-date-picker/nz-picker/span/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/calendar-footer/div/span/today-button/a').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[13]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-select/div/div').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[5]/div/div/div/ul/li[4]').click()
        time.sleep(1)
        # 12.一级名录，产品名录
        driver.find_element(By.XPATH,
                            '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[6]/div/div/div/div/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,
                            '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[11]/td[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH,
                            '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        time.sleep(2)

        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[5]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[2]/td[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        time.sleep(2)
        # 13.业务承接所属
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[7]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-input-group/span/span[2]').click()
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-input-group/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH,'/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[2]').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[5]/div/div/div[3]/button[1]').click()
        # 14.商机重要程度
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[10]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-select/div/div').click()
        time.sleep(2)
        print(18)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[5]/div/div/div/ul/li[3]').click()
        time.sleep(2)
        # 15.预计签约时间
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[11]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-date-picker/nz-picker/span/input').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[5]/div/div/date-range-popup/div/div/div/div/calendar-footer/div/span/today-button/a').click()
        time.sleep(2)
        # 16.客户联系人、电话
        driver.find_element(By.XPATH,'//*[@id="test1574041790960110b5b726c"]').click()
        time.sleep(2)
        driver.find_element(By.XPATH,'//*[@id="test1574041790960110b5b726c"]').send_keys('/')
        time.sleep(2)
        driver.find_element(By.XPATH, '//*[@id="test1628832276753244976d1e2"]').click()
        time.sleep(2)
        driver.find_element(By.XPATH, '//*[@id="test1628832276753244976d1e2"]').send_keys('/')
        time.sleep(2)
        # 17.保存
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/form-table/form-preview/pop-window[1]/div/div/div[2]/div[3]/div[2]/nz-form-control/div/span/button[1]').click()
        time.sleep(2)
        print('finish')
        time.sleep(5)
        driver.close()
        return True
    except Exception:
        return False


def read_excel(main_excel_name):
    a = []
    workbook = openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = workbook.active
    for i in sheet.iter_rows():
        for j in i:
            a.append(j.value)
    return a


def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')


def yyz_main(main_excel_name):
    # 把表格名拼接出日志文件名
    log_file_name = 'log_' + main_excel_name + '.txt'
    # 建空列表记录运行错误的编号
    error_data = []
    # 读excel
    wb = openpyxl.load_workbook(main_excel_name + '.xlsx')
    # 默认sheet，如果指定sheet名就 sheet = wb.active('sheet名')
    sheet = wb.active
    # 因为第一行是标题，获取最大长度要-1，比如10行的excel就是9个数据,test_num=9
    test_num = sheet.max_row - 1
    # 从2开始遍历到9+2，因为python是左闭右开，所以代码是[2,11),从第二行读到第十行,一共9次
    for i in range(2, test_num + 2):
        # 获取对应行数和第1列的项目编号，和第2列的工作量
        project_num = sheet.cell(row=i, column=1).value
        workload = sheet.cell(row=i, column=2).value
        col3_data = sheet.cell(row=i, column=3).value
        # 打日志
        print('共{}条数据'.format(test_num))
        start_log = 'No:{},project_num:{},workload:{},col3_data:{},start_time:{}'.format(i - 1, project_num, workload,col3_data,time.strftime('%Y-%m-%d %H:%M:%S'))
        print(start_log)
        write_log_data(log_file_name, start_log)
        # 运行selenium脚本，把这次运行的project_num和workload传进去，返回类型为布尔型，True或者False
        flag = main_point(project_num, workload, col3_data)
        # 按返回类型打日志是否运行成功
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            # 把运行错误的编号存到error_data列表
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')
    # 那列表长度确定错误数
    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(test_num,
                                                                                                 test_num - fail_data_sum,
                                                                                                 fail_data_sum)
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, i)


if __name__ == '__main__':
    # 输入获取表格名，传到主函数
    excel_name = input('表格名：')
    yyz_main(excel_name)
