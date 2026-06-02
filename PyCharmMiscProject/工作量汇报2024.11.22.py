from datetime import datetime
import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def change_col3_data(driver, i, col3_data):
    time.sleep(1)
    driver.find_element(By.XPATH,
                        '//*[@id="body"]/div[2]/div/form/div[17]/div/div/div/div/div/nz-tabset/div[2]/div[1]/div/div/div/div/div/div/button').click()
    # 17，点击空白框
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr[{}]/td[2]/div/div'.format(i)).click()
    time.sleep(1)
    # 18，点击编辑图标
    driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr[{}]/td[2]/nz-input-group/span/span[2]'.format(i)).click()
    time.sleep(1)
    # 19，选择人
    time.sleep(2)
    page_i = i
    while page_i > 30:
        page_i = page_i - 30
        print(page_i)
    worker_element = driver.find_element(By.XPATH,
                                         '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[{}]/td[1]'.format(
        page_i))
    ActionChains(driver).double_click(worker_element).perform()
    time.sleep(1)
    # 20，点击合计工时框
    driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr[{}]/td[5]/div/div'.format(i)).click()
    # 21，输入工时
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr[{}]/td[5]/nz-input-number/div[2]/input'.format(i)).send_keys(col3_data)


def main_point(project_num, workload, col3_data):
    """
    :param project_num:项目编号
    :param workload:需汇报工作量
    :param col3_data:人工工时
    :return: 布尔型，返回运行结果
    """
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://bpm.chinajyy.net/#/passport/login')
        driver.maximize_window()
        driver.find_element(By.CLASS_NAME, 'ant-input').send_keys('Z0343')
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button').click()
        # 3，自建应用
        time.sleep(5)
        el0 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span')))
        el0.click()
        # 4，项目管理
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-1"]/div/div/div/nz-spin/div/div[1]/div[4]/small').click()
        # 6，查询方案
        time.sleep(5)
        el1 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[1]/div[1]/button[1]')))
        el1.click()
        # 7，加号图标
        time.sleep(3)
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[1]/div/div/a[1]').click()
        # driver.find_element(By.CLASS_NAME,'ant-btn-default').click()
        # 8，【过滤条件：项目编码，比较符：等于，比较值：项目编码】
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[3]/nz-select/div').click()
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-3"]/div/div/ul/li[1]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[4]/nz-select/div/div').click()
        driver.find_element(By.XPATH, '//*[@id="cdk-overlay-4"]/div/div/ul/li[1]').click()
        driver.find_element(By.XPATH, '//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[2]/div/div[2]/table/tbody/tr/td[5]/input').send_keys(project_num)
        # 9，确定按钮
        driver.find_element(By.XPATH, '//*[@id="footer"]/div/div/button[1]').click()
        # 10，选择项目编码对应项目
        time.sleep(6)
        el2 = WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr/td[2]')))
        el2.click()
        # 11，关联生成
        driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[1]/div[1]/nz-dropdown[1]/button').click()
        time.sleep(1)
        # 12，生成项目工作汇报单
        driver.find_element(By.XPATH, '/html/body/div[3]/div[4]/div/div/div/ul[3]/li/div').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div[3]/div[5]/div/div/ul/ul[1]/li/a').click()
        # 16，新增数据
        time.sleep(1)
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[16]/div/div/div/div/div/div/div/nz-form-control/div/span/textarea").click()
        time.sleep(1)
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[16]/div/div/div/div/div/div/div/nz-form-control/div/span/textarea").send_keys("项目在建中")
        # 新增一条数据
        driver.find_element(By.XPATH, "/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[17]/div/div/div/div/div/nz-tabset/div[2]/div[1]/div/div/div/div/div/div/button").click()
        # 17，点击空白框
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr/td[2]/div/div').click()
        time.sleep(1)
        # 18，点击编辑图标
        driver.find_element(By.XPATH, '//*[@id="childTableBody-ProWorkReportEW"]/tr/td[2]/nz-input-group/span/span[2]').click()
        # 检验成员表人数
        time.sleep(2)
        person_num_str = driver.find_element(By.XPATH, '//*[@id="body"]/div/nz-pagination/ul/li[1]').text
        person_num = int(person_num_str[1:-1])
        print(person_num)
        time.sleep(1)
        # 19，选择人
        time.sleep(3)
        worker_element = driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/nz-table/nz-spin/div/div/div/div/div[2]/table/tbody/tr[1]/td[1]')
        ActionChains(driver).double_click(worker_element).perform()
        time.sleep(1)
        # 执行多次工人的工时输入动作
        for i in range(1, person_num + 1):
            change_col3_data(driver, i, col3_data)
        if 'i' > 30 and 'i' < 60:
            # 人数超过30点击两次下一页
            driver.find_element(By.XPATH,
                                '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/div/nz-pagination/ul/li[6]/a/i/svg').click()
            time.sleep(2)
        elif 'i' > 60:
            # 人数超过60点击一次下一页
            driver.find_element(By.XPATH,
                                '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/div/nz-pagination/ul/li[6]/a/i/svg').click()
            time.sleep(2)
            driver.find_element(By.XPATH,
                                '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[5]/div/div/div[2]/div/nz-pagination/ul/li[6]/a/i/svg').click()
        # 22，勾选”工作量调整，允许跳转累计汇报工作量“
        driver.find_element(By.XPATH,
                            '//*[@id="body"]/div[2]/div/form/div[17]/div/div/div/div/div/nz-tabset/div[2]/div[1]/div/div/div/div/div/div/button').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '//*[@id="body"]/div[2]/div/form/div[7]/div/div/div/div[2]/div/div/div/nz-form-control/div/span/nz-checkbox-wrapper/div/div/label/span[1]/input').click()
        time.sleep(1)
        # 23，本期汇报金额为workload
        driver.find_element(By.XPATH, "//*[@id='body']/div[2]/div/form/div[11]/div/div/div/div[1]/div/div/div/nz-form-control/div/span/nz-input-number/div[2]/input").send_keys(workload)
        time.sleep(1)
        el = driver.find_element(By.XPATH, '/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/form-preview/pop-window[1]/div/div/div[2]/div[2]/div/form/div[20]/div/div/div/nz-form-control/div/span/div[1]/div[1]/div/nz-upload/div/div/input')
        el.send_keys(r"C:\Users\admin\PyCharmMiscProject\2025年8月工作量签单.pdf")
        time.sleep(5)
        # 23，审核ZCHJC003869
        driver.find_element(By.XPATH, '//*[@id="body"]/div[3]/div/nz-form-control/div/span/button[3]').click()
        print('finish')
        time.sleep(5)
        driver.close()
        return True
    except Exception:
        return False


def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')


def yyz_main(main_excel_name):
    log_file_name = 'log_' + main_excel_name + '.txt'
    error_data = []
    wb = openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = wb.active
    test_num = sheet.max_row - 1
    for i in range(2, test_num + 2):
        project_num = sheet.cell(row=i, column=1).value
        workload = sheet.cell(row=i, column=2).value
        col3_data = sheet.cell(row=i, column=3).value
        print('共{}条数据'.format(test_num))
        start_log = 'No:{},project_num:{},workload:{},start_time:{}'.format(i - 1, project_num, workload, time.strftime('%Y-%m-%d %H:%M:%S'))
        print(start_log)
        write_log_data(log_file_name, start_log)
        flag = main_point(project_num, workload, col3_data)
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')
    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(test_num, test_num - fail_data_sum, fail_data_sum)
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, i)


if __name__ == '__main__':
    excel_name = input('表格名：')
    yyz_main(excel_name)
