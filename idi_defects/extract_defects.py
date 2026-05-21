"""IDI 质量险潜在缺陷清单 - 结构化提取器

从 docx 中读取 20 个表格,把每条缺陷条目转成结构化数据,
输出 defects.json 和 defects.xlsx,供检索/报告生成模块使用。
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path

import docx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# 表格序号 -> (风险期限, 大类) 映射,按 docx 目录顺序硬编码
TABLE_CATEGORY = {
    0:  ("十年期", "地基基础工程"),
    1:  ("十年期", "主体结构工程"),
    2:  ("五年期", "防水工程"),
    3:  ("五年期", "保温工程"),
    4:  ("两年期", "砌体工程"),
    5:  ("两年期", "抹灰工程"),
    6:  ("两年期", "涂饰、软包工程"),
    7:  ("两年期", "吊顶工程"),
    8:  ("两年期", "墙面铺装与隔墙工程"),
    9:  ("两年期", "地面铺装工程"),
    10: ("两年期", "门窗工程"),
    11: ("两年期", "栏杆、楼梯工程"),
    12: ("两年期", "幕墙工程"),
    13: ("两年期", "细部工程"),
    14: ("两年期", "其他"),
    15: ("机电安装", "通风与空调安装工程"),
    16: ("机电安装", "电气安装工程"),
    17: ("机电安装", "给排水及消防管道工程"),
    18: ("机电安装", "智能化工程"),
    19: ("机电安装", "电梯工程"),
}


@dataclass
class Defect:
    """单条缺陷条目"""

    category_period: str   # 风险期限:十年期/五年期/两年期/机电安装
    category_major: str    # 大类工程
    category_minor: str    # 子分类(分项标题,如"桩基工程1:管桩")
    seq: str               # 表内序号
    description: str       # 完整缺陷描述原文
    problem: str           # 提取出的"问题描述"主体
    standard: str          # 提取出的"违反/不符合"的规范条文
    suggestion: str        # 纠正和预防建议
    submitter: str         # 提交人/备忘
    global_id: str = ""    # 全局唯一 ID,extract 时填


# ---------- 文本解析辅助 ----------

_NORMALIZE_SPACE = re.compile(r"[\s　]+")


def norm(text: str) -> str:
    """规范化空白:把连续空白/全角空格合并为一个空格,首尾去空白"""
    return _NORMALIZE_SPACE.sub(" ", text or "").strip()


def split_problem_standard(desc: str) -> tuple[str, str]:
    """从缺陷描述里拆出"问题描述主体"和"违反的规范条文"两部分。

    模板大约是:"问题描述:XXXXXX。违反YYY规范ZZ条的要求。"
    如果没有"违反/不符合"关键词,则全部作为问题主体返回。
    """
    text = norm(desc)
    if not text:
        return "", ""

    # 去掉开头的"问题描述:"前缀
    text_body = re.sub(r"^问题描述[::]\s*", "", text)

    # 找到第一个"违反"或"不符合"的位置作为分割点
    m = re.search(r"(违反|不符合)", text_body)
    if not m:
        return text_body, ""

    problem = text_body[: m.start()].rstrip("。.,;; ")
    standard = text_body[m.start():].strip()
    return problem, standard


def is_section_header_row(cells: list[str]) -> bool:
    """子分项标题行的特征:同一行所有 cell 文字相同且非空(merge cell 展开后效果)。"""
    nonempty = [c for c in cells if c]
    if len(nonempty) < 2:
        return False
    return all(c == nonempty[0] for c in nonempty)


def is_data_row(cells: list[str]) -> bool:
    """数据行:第一列是数字(序号),第3/4列(描述/建议)至少一个非空"""
    if not cells:
        return False
    seq = cells[0].strip()
    if not re.fullmatch(r"\d+", seq):
        return False
    desc = cells[2] if len(cells) > 2 else ""
    sug = cells[3] if len(cells) > 3 else ""
    return bool(desc or sug)


# ---------- 主提取流程 ----------

def extract(docx_path: Path) -> list[Defect]:
    doc = docx.Document(str(docx_path))
    results: list[Defect] = []

    for ti, table in enumerate(doc.tables):
        if ti not in TABLE_CATEGORY:
            continue
        period, major = TABLE_CATEGORY[ti]

        # 当前子分类,从分项标题行更新;若没有则用大类名兜底
        current_minor = major

        # 跳过表头行(行 0):列名"序号/问题图片/缺陷描述/纠正和预防建议/备忘"
        for ri, row in enumerate(table.rows):
            if ri == 0:
                continue
            cells = [norm(c.text) for c in row.cells]

            # 分项标题行(同行 merged 全相同)
            if is_section_header_row(cells):
                current_minor = cells[0]
                continue

            # 数据行
            if not is_data_row(cells):
                continue

            seq = cells[0]
            desc = cells[2] if len(cells) > 2 else ""
            sug = cells[3] if len(cells) > 3 else ""
            submitter = cells[4] if len(cells) > 4 else ""

            problem, standard = split_problem_standard(desc)

            results.append(
                Defect(
                    category_period=period,
                    category_major=major,
                    category_minor=current_minor,
                    seq=seq,
                    description=desc,
                    problem=problem,
                    standard=standard,
                    suggestion=sug,
                    submitter=submitter,
                )
            )

    # 全局 ID
    for i, d in enumerate(results, start=1):
        d.global_id = f"D{i:04d}"

    return results


# ---------- 导出 ----------

def export_json(defects: list[Defect], out_path: Path) -> None:
    data = [asdict(d) for d in defects]
    out_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def export_excel(defects: list[Defect], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "IDI缺陷清单"

    headers = [
        "全局ID", "风险期限", "大类工程", "子分类",
        "原序号", "问题描述", "违反规范", "纠正和预防建议",
        "提交人/备忘", "原始缺陷描述(完整)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for d in defects:
        ws.append([
            d.global_id, d.category_period, d.category_major, d.category_minor,
            d.seq, d.problem, d.standard, d.suggestion,
            d.submitter, d.description,
        ])

    # 列宽
    widths = [10, 10, 16, 22, 8, 50, 50, 50, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # 数据行换行
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ---------- CLI ----------

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="从 IDI 缺陷清单 docx 提取结构化数据")
    parser.add_argument(
        "--input", "-i",
        default=r"C:\Users\admin\xwechat_files\wxid_6kuqjhsipy5321_aa45\msg\file\2026-05\IDI质量险潜在缺陷清单合集 2025.03.22.docx",
        help="源 docx 路径",
    )
    parser.add_argument(
        "--outdir", "-o",
        default=str(Path(__file__).parent / "data"),
        help="输出目录",
    )
    args = parser.parse_args()

    docx_path = Path(args.input)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    print(f"读取: {docx_path}")
    defects = extract(docx_path)
    print(f"提取条目数: {len(defects)}")

    # 按大类统计
    from collections import Counter
    cnt = Counter((d.category_period, d.category_major) for d in defects)
    print("\n按类别统计:")
    for (period, major), n in cnt.items():
        print(f"  [{period}] {major}: {n} 条")

    json_path = outdir / "defects.json"
    xlsx_path = outdir / "defects.xlsx"
    export_json(defects, json_path)
    export_excel(defects, xlsx_path)
    print(f"\n已输出:\n  {json_path}\n  {xlsx_path}")


if __name__ == "__main__":
    main()
