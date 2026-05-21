#!/usr/bin/env python3
"""PDF 拆分 / 合并 / 批量重命名 / 压缩 / 报告 工具"""

import argparse
import glob
import os
import re
import sys
from datetime import datetime


def human_size(size_bytes):
    """将字节转为人类可读格式"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if abs(size_bytes) < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} TB"


def split_pdf(pdf_path, output_dir="split"):
    """按页拆分 PDF"""
    import pikepdf

    if not os.path.exists(pdf_path):
        print(f"[X] 文件不存在: {pdf_path}")
        return

    os.makedirs(output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(pdf_path))[0]

    try:
        pdf = pikepdf.open(pdf_path)
        count = len(pdf.pages)

        for i, page in enumerate(pdf.pages, 1):
            out = pikepdf.Pdf.new()
            out.pages.append(page)
            out_path = os.path.join(output_dir, f"{base}_page_{i:03d}.pdf")
            out.save(out_path)
            print(f"  [FILE] {out_path}")

        print(f"[OK] 拆分完成: {count} 页 → {output_dir}/")
        pdf.close()

    except Exception as e:
        print(f"[X] 拆分失败: {e}")


def generate_txt_report(report_data, output_path):
    """生成文本报告"""
    lines = []
    lines.append("=" * 60)
    lines.append("  PDF 合并报告")
    lines.append("=" * 60)
    lines.append(f"生成时间: {report_data['timestamp']}")
    lines.append(f"输出文件: {report_data['output_file']}")
    lines.append("")
    lines.append(f"{'序号':<6}{'文件名':<30}{'页数':<8}{'大小':<12}")
    lines.append("-" * 60)

    for i, item in enumerate(report_data['files'], 1):
        lines.append(f"{i:<6}{item['name']:<30}{item['pages']:<8}{item['size']:<12}")

    lines.append("-" * 60)
    lines.append(f"{'合计':<6}{report_data['total_files']} 个文件{'':<17}{report_data['total_pages']} 页{'':<6}{report_data['total_size']}")
    lines.append("=" * 60)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"[OK] 文本报告已生成: {output_path}")


def generate_excel_report(report_data, output_path):
    """生成 Excel 报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("[!] 未安装 openpyxl，跳过 Excel 报告生成")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "合并清单"

    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = "PDF 合并报告"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # 信息行
    ws['A2'] = f"生成时间: {report_data['timestamp']}"
    ws['A3'] = f"输出文件: {report_data['output_file']}"
    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')

    # 表头
    headers = ['序号', '文件名', '页数', '文件大小']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for i, item in enumerate(report_data['files'], 1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=item['name']).border = thin_border
        ws.cell(row=row, column=3, value=item['pages']).border = thin_border
        ws.cell(row=row, column=4, value=item['size']).border = thin_border

    # 合计行
    total_row = 5 + len(report_data['files']) + 1
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"{report_data['total_files']} 个文件").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=report_data['total_pages']).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=report_data['total_size']).font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).border = thin_border

    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15

    wb.save(output_path)
    print(f"[OK] Excel 报告已生成: {output_path}")


def merge_pdf(pdf_list, output_path, report=False):
    """合并多个 PDF，可选生成报告"""
    import pikepdf

    if not pdf_list:
        print("[X] 未提供输入文件")
        return

    # 展开通配符
    expanded = []
    for p in pdf_list:
        if '*' in p or '?' in p:
            expanded.extend(glob.glob(p))
        else:
            expanded.append(p)

    missing = [f for f in expanded if not os.path.exists(f)]
    if missing:
        print(f"[X] 文件不存在: {', '.join(missing)}")
        return

    report_data = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'output_file': os.path.basename(output_path),
        'files': [],
        'total_files': 0,
        'total_pages': 0,
        'total_size': 0,
    }

    try:
        merged = pikepdf.Pdf.new()
        total_pages = 0
        total_size_bytes = 0

        for f in expanded:
            src = pikepdf.open(f)
            file_size = os.path.getsize(f)
            total_size_bytes += file_size
            page_count = len(src.pages)
            total_pages += page_count

            report_data['files'].append({
                'name': os.path.basename(f),
                'pages': page_count,
                'size': human_size(file_size),
                'size_bytes': file_size,
            })

            merged.pages.extend(src.pages)
            src.close()
            print(f"  [+] {f} ({page_count} 页)")

        merged.save(output_path)
        merged.close()

        output_size = os.path.getsize(output_path)
        report_data['total_files'] = len(expanded)
        report_data['total_pages'] = total_pages
        report_data['total_size'] = human_size(total_size_bytes)

        print(f"[OK] 合并完成: {total_pages} 页 → {output_path}")
        print(f"     输出大小: {human_size(output_size)}")

        # 生成报告
        if report:
            base = os.path.splitext(output_path)[0]
            txt_path = base + "_报告.txt"
            xlsx_path = base + "_报告.xlsx"
            generate_txt_report(report_data, txt_path)
            generate_excel_report(report_data, xlsx_path)

    except Exception as e:
        print(f"[X] 合并失败: {e}")


def rename_pdfs(pattern, replacement, folder=".", preview=False):
    """批量重命名 PDF（支持正则）"""
    files = glob.glob(os.path.join(folder, "*.pdf"))
    if not files:
        print(f"[!]  目录中无 PDF 文件: {folder}")
        return

    changes = []
    for f in sorted(files):
        old_name = os.path.basename(f)
        new_name = re.sub(pattern, replacement, old_name)

        if new_name != old_name:
            old_path = f
            new_path = os.path.join(folder, new_name)
            changes.append((old_path, new_path, old_name, new_name))

    if not changes:
        print("[!]  无文件需要重命名")
        return

    print(f"[LIST] 将重命名 {len(changes)} 个文件:")
    for _, _, old, new in changes:
        print(f"   {old} → {new}")

    if preview:
        print("\n[EYE] 预览模式，未执行重命名")
        return

    confirm = input("\n确认执行? [y/N]: ")
    if confirm.lower() != 'y':
        print("[X] 已取消")
        return

    for old_path, new_path, old, new in changes:
        try:
            os.rename(old_path, new_path)
            print(f"[OK] {old} → {new}")
        except Exception as e:
            print(f"[X] 失败 {old}: {e}")


def extract_pages(pdf_path, pages, output_path):
    """提取指定页码（如 1,3,5-10）"""
    import pikepdf

    def parse_pages(spec):
        result = []
        for part in spec.split(','):
            if '-' in part:
                start, end = part.split('-')
                result.extend(range(int(start), int(end) + 1))
            else:
                result.append(int(part))
        return result

    try:
        pdf = pikepdf.open(pdf_path)
        page_nums = parse_pages(pages)

        out = pikepdf.Pdf.new()
        for n in page_nums:
            if 1 <= n <= len(pdf.pages):
                out.pages.append(pdf.pages[n - 1])

        out.save(output_path)
        out.close()
        pdf.close()
        print(f"[OK] 提取 {len(page_nums)} 页 → {output_path}")

    except Exception as e:
        print(f"[X] 提取失败: {e}")


def compress_pdf(input_path, output_path=None, image_quality="medium"):
    """压缩 PDF 文件"""
    import pikepdf
    from io import BytesIO

    if not os.path.exists(input_path):
        print(f"[X] 文件不存在: {input_path}")
        return

    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_压缩{ext}"

    quality_map = {
        "low": 30,
        "medium": 50,
        "high": 75,
    }
    jpeg_quality = quality_map.get(image_quality, 50)

    original_size = os.path.getsize(input_path)
    print(f"[*] 原始大小: {human_size(original_size)}")
    print(f"[*] 压缩模式: {image_quality} (JPEG 质量 {jpeg_quality})")

    try:
        pdf = pikepdf.open(input_path)

        # 尝试用 Pillow 压缩图片
        try:
            from PIL import Image
            images_compressed = 0

            for page in pdf.pages:
                if "/Resources" not in page:
                    continue
                resources = page.Resources
                if "/XObject" not in resources:
                    continue

                xobjects = resources.XObject
                for name, xobj in list(xobjects.items()):
                    try:
                        obj = pikepdf.Pdf.open(xobj.objgen)
                        if obj.get("/Subtype") != "/Image":
                            continue

                        width = int(obj.get("/Width", 0))
                        height = int(obj.get("/Height", 0))
                        if width == 0 or height == 0:
                            continue

                        filter_type = obj.get("/Filter")
                        if filter_type == "/DCTDecode":
                            # 已经是 JPEG，尝试降低质量
                            img_data = obj.read_raw_bytes()
                            img = Image.open(BytesIO(img_data))
                            buf = BytesIO()
                            img.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
                            obj.read_bytes = buf.getvalue()
                            images_compressed += 1
                        elif filter_type == "/FlateDecode":
                            # PNG 类图片，转为 JPEG 压缩
                            color_space = obj.get("/ColorSpace", "/DeviceRGB")
                            if color_space == "/DeviceRGB":
                                mode = "RGB"
                            elif color_space == "/DeviceGray":
                                mode = "L"
                            else:
                                continue

                            img_data = obj.read_raw_bytes()
                            img = Image.frombytes(mode, (width, height), img_data)
                            buf = BytesIO()
                            img.save(buf, format="JPEG", quality=jpeg_quality, optimize=True)
                            obj.read_bytes = buf.getvalue()
                            obj.Filter = pikepdf.Name("/DCTDecode")
                            if "/DecodeParms" in obj:
                                del obj.DecodeParms
                            images_compressed += 1
                    except Exception:
                        continue

            if images_compressed > 0:
                print(f"[*] 压缩了 {images_compressed} 张图片")

        except ImportError:
            print("[!] 未安装 Pillow，跳过图片压缩")

        # 保存（线性化优化）
        pdf.save(output_path, linearize=True)
        pdf.close()

        new_size = os.path.getsize(output_path)
        ratio = (1 - new_size / original_size) * 100 if original_size > 0 else 0
        print(f"[OK] 压缩完成: {human_size(original_size)} → {human_size(new_size)} (减少 {ratio:.1f}%)")

    except Exception as e:
        print(f"[X] 压缩失败: {e}")


def main():
    parser = argparse.ArgumentParser(description="PDF 工具集")
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # split
    p_split = subparsers.add_parser("split", help="按页拆分PDF")
    p_split.add_argument("input", help="输入PDF")
    p_split.add_argument("-o", "--output", default="split", help="输出目录")

    # merge
    p_merge = subparsers.add_parser("merge", help="合并PDF")
    p_merge.add_argument("inputs", nargs="+", help="输入PDF文件")
    p_merge.add_argument("-o", "--output", required=True, help="输出PDF")
    p_merge.add_argument("-r", "--report", action="store_true", help="生成合并报告")

    # rename
    p_rename = subparsers.add_parser("rename", help="批量重命名")
    p_rename.add_argument("pattern", help="匹配正则")
    p_rename.add_argument("replacement", help="替换字符串")
    p_rename.add_argument("-d", "--dir", default=".", help="目标目录")
    p_rename.add_argument("-n", "--dry-run", action="store_true", help="预览模式")

    # extract
    p_extract = subparsers.add_parser("extract", help="提取指定页")
    p_extract.add_argument("input", help="输入PDF")
    p_extract.add_argument("pages", help="页码（如 1,3,5-10）")
    p_extract.add_argument("-o", "--output", required=True, help="输出PDF")

    # compress
    p_compress = subparsers.add_parser("compress", help="压缩PDF")
    p_compress.add_argument("input", help="输入PDF")
    p_compress.add_argument("-o", "--output", help="输出PDF（默认原文件名_压缩.pdf）")
    p_compress.add_argument("-q", "--quality", choices=["low", "medium", "high"],
                            default="medium", help="压缩质量 (默认 medium)")

    args = parser.parse_args()

    if args.command == "split":
        split_pdf(args.input, args.output)
    elif args.command == "merge":
        merge_pdf(args.inputs, args.output, report=args.report)
    elif args.command == "rename":
        rename_pdfs(args.pattern, args.replacement, args.dir, args.dry_run)
    elif args.command == "extract":
        extract_pages(args.input, args.pages, args.output)
    elif args.command == "compress":
        compress_pdf(args.input, args.output, args.quality)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
