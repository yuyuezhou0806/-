#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
填写合同清单报价表J列
J列 = J列合并单元格所覆盖的那些行对应的I列数值之和
"""

import openpyxl
from openpyxl import load_workbook
import os


def is_number(value):
    """检查是否为有效数字"""
    if value is None or value == '-' or value == '':
        return False
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def get_cell_value(sheet_data, row, col):
    """获取单元格的计算值（使用data_only模式加载的工作簿）"""
    try:
        cell = sheet_data.cell(row=row, column=col)
        return cell.value
    except:
        return None


def get_j_merged_ranges(sheet):
    """获取所有J列（第10列）的合并单元格范围"""
    j_ranges = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_col <= 10 <= merged_range.max_col:
            j_ranges.append({
                'start_row': merged_range.min_row,
                'end_row': merged_range.max_row,
                'main_cell_row': merged_range.min_row  # 主单元格所在行
            })
    # 按起始行排序
    j_ranges.sort(key=lambda x: x['start_row'])
    return j_ranges


def process_excel(filepath):
    """处理Excel文件，填写J列"""

    if not os.path.exists(filepath):
        print(f"错误：文件不存在 {filepath}")
        return False

    print(f"正在加载文件: {filepath}")

    # 用data_only模式读取计算后的值
    wb_data = load_workbook(filepath, data_only=True)
    sheet_data = wb_data.active

    # 用普通模式打开用于写入
    wb = load_workbook(filepath)
    sheet = wb.active

    print(f"工作表: {sheet.title}, 总行数: {sheet.max_row}")

    # 获取J列所有合并单元格范围
    j_ranges = get_j_merged_ranges(sheet)
    print(f"J列合并单元格数量: {len(j_ranges)}")

    # 处理每个J列合并单元格
    print("\n正在计算并填写J列...")

    for idx, j_range in enumerate(j_ranges, 1):
        start_row = j_range['start_row']
        end_row = j_range['end_row']
        main_row = j_range['main_cell_row']

        # 只处理数据行（从第5行开始）
        if start_row < 5:
            continue

        # 计算该合并单元格范围内所有I列的值之和
        i_sum = 0
        i_details = []

        for row in range(start_row, end_row + 1):
            i_value = get_cell_value(sheet_data, row, 9)  # I列是第9列
            if is_number(i_value):
                i_num = float(i_value)
                i_sum += i_num
                i_details.append(f"I{row}={i_num}")

        print(f"  J{start_row}-J{end_row} (主单元格J{main_row}): {', '.join(i_details)} -> 总和={i_sum}")

        # 填写到J列的主单元格
        main_cell = sheet.cell(row=main_row, column=10)  # J列是第10列
        main_cell.value = i_sum

    # 保存文件
    output_path = filepath.replace('.xlsx', '_已填写.xlsx')
    wb.save(output_path)
    print(f"\n完成！已保存到: {output_path}")

    # 关闭workbook
    wb_data.close()
    wb.close()

    return True


if __name__ == '__main__':
    # 文件路径
    filepath = r"C:\Users\admin\Desktop\附件4：合同清单报价-标段一-zch111.xlsx"

    process_excel(filepath)
