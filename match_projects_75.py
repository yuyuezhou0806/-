import openpyxl
from difflib import SequenceMatcher
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Data from screenshot: (客户名称, 工程名称)
screenshot_data = [
    ("上海沪申高速公路建设发展有限公司", "浦东国际机场T3航站楼S32公路立交改建工程施工2标（主体）"),
    ("上海市浦东新区道路运输事业发展中心", "S20(杨高南路立交-上南路)上行道路及排水管道维修工程施工检测"),
    ("上海中咨安通工程管理股份有限公司", "S225朱枫公路(K9+950-K18+088.8、K24+800-K26+709)修复养护工程"),
    ("上海宝山公路工程建设有限公司", "闸航路（三鲁公路-浦星公路）修复养护工程"),
    ("上海浦桥工程建设管理有限公司", "S4公路入城段（莘庄立交-金都路）交通功能完善工程-二标段春申塘以北段"),
    ("上海市道路运输事业发展中心", "2025年部分市管公路路面修复养护工程（9标）"),
    ("上海市奉贤区交通建设管理中心", "望园路（金齐路~大叶公路）道路改建工程"),
    ("太仓市路桥工程有限公司", "太仓融创城市建设发展有限公司改造双凤镇凤杨路FYL-SG标段"),
    ("上海市浦东新区道路运输事业发展中心", "S20(杨高南路立交-上南路)上行道路及排水管道维修工程"),
    ("上海市青浦区人民政府香花桥街道办事处", "2025年农村公路大中修项目香花桥街道清河湾路（漕盈路-西大盈港）整治工程"),
    ("中铁二局集团有限公司暨南大道西延惠山段（江阴界~S261）新建工程项目经理部", "暨南大道西延惠山段（江阴界～S261）新建工程"),
    ("上海公路工程监理有限公司", "叶新-大叶公路（叶榭镇界-奉贤区界）改建工程04标"),
    ("上海市奉贤区交通建设管理中心", "奉贤区望园路（金齐路~大叶公路）道路改建工程1标"),
    ("上海浦桥工程建设管理有限公司", "S4公路入城段（莘庄立交-金都路）交通功能完善工程春申塘桥以南段（三标段）"),
    ("上海市青浦区道路运输管理事务中心", "新太公路（沪苏省界-青赵公路）新建工程"),
    ("上海浦桥工程建设管理有限公司", "S4公路入城段（莘庄立交-金都路）交通功能完善工程"),
    ("上海东华地方铁路开发有限公司", "南陈路-南秀路拓宽改建工程一标"),
    ("上海城投航道建设有限公司", "苏申内港线暨吴淞江（省界-老白石路）整治工程施工3标"),
    ("上海同结建设工程有限公司", "市管公路交通安全设施品质提升专项养护工程（一期）5标"),
    ("上海市青浦区赵巷镇人民政府", "2025年农村公路大中修项目赵巷镇佳悦路（盈港东路-佳康路）"),
    ("上海浦桥工程建设管理有限公司", "S4公路入城段（莘庄立交-金都路）交通功能完善工程春申塘以南段（三标段）"),
    ("上海华新城镇建设管理有限公司", "青浦区华新镇华志路（芦蔡北路-华杨河）新建工程"),
    ("练塘镇金前村村民委员会", "2025年度金前村农村公益事业“一事一议”项目"),
    ("练塘镇北埭村村民委员会", "2025年度北埭村农村公益事业“一事一议”项目"),
    ("苏州市路达工程监理咨询有限公司", "东太湖隧道连接线（旺山路）工程检测WSL-JC1标段"),
    ("上海建工四建集团有限公司", "香花桥街道农村公路二类专项维修工程"),
    ("上海市道路运输事业发展中心", "S322宝安线（K0+000～K15+010）修复养护工程6标段"),
    ("上海嘉众市政建设工程有限公司", "S322 宝安线（K0+000～K15+010）修复养护工程5标"),
    ("上海市道路运输事业发展中心", "S322宝安线（K0+000-K15+010）修复养护工程三标段"),
    ("上海市闵行区交通运输事业发展中心", "华西路（金都路-颛兴路）道路大修工程"),
    ("上海建工四建集团有限公司", "香花桥街道天辰路（华青路-久乐路）小修工程"),
    ("上海市金山区漕泾镇人民政府", "漕泾镇朱漕公路（G228公路-已建段）新建工程与G228公路-金光路增设道口工程"),
    ("上海市金山区漕泾镇人民政府", "漕泾镇朱漕公路（G228公路-已建段）新建工程"),
]

print("Reading Excel...")
wb = openpyxl.load_workbook(r'C:\Users\admin\Desktop\htgl2026.xlsx')
ws = wb.active

excel_rows = []
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    proj = str(row[12] or "").strip()
    client = str(row[13] or "").strip()
    note = str(row[2] or "").strip()
    archive = str(row[34] or "").strip()
    if proj and proj not in ('/', '\\'):
        excel_rows.append({
            "row": idx,
            "proj": proj,
            "client": client,
            "note": note,
            "archive": archive
        })

print(f"Loaded {len(excel_rows)} rows from Excel")

results = []
for s_client, s_proj in screenshot_data:
    best_match = None
    best_score = 0
    for e in excel_rows:
        score = similarity(s_proj, e["proj"])
        if score > best_score:
            best_score = score
            best_match = e

    matched = best_score >= 0.75
    results.append({
        "工程名称": s_proj,
        "客户名称": s_client,
        "合同编号备注": best_match["note"] if matched else "",
        "归档状态": best_match["archive"] if matched else "未匹配",
        "匹配到的Excel工程名称": best_match["proj"] if not matched else "",
        "相似度": round(best_score, 4),
        "是否匹配": matched
    })
    status = "MATCH" if matched else "SKIP"
    print(f"[{status}] {s_proj[:40]}... -> {best_score:.2%}")

# Write output Excel
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "匹配结果"

headers = ["工程名称", "客户名称", "合同编号备注", "归档状态", "匹配到的Excel工程名称", "相似度"]
out_ws.append(headers)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in out_ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Fill data
for r in results:
    row_data = [
        r["工程名称"],
        r["客户名称"],
        r["合同编号备注"],
        r["归档状态"],
        r["匹配到的Excel工程名称"],
        f"{r['相似度']:.2%}"
    ]
    out_ws.append(row_data)

# Style and column widths
for col in out_ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    out_ws.column_dimensions[col_letter].width = adjusted_width

out_ws.row_dimensions[1].height = 25

# Freeze header
out_ws.freeze_panes = 'A2'

output_path = r'C:\Users\admin\Desktop\工程匹配结果.xlsx'
out_wb.save(output_path)
print(f"\nOutput saved to: {output_path}")

matched_count = sum(1 for r in results if r["是否匹配"])
print(f"Summary: {matched_count}/{len(results)} matched (>=90% similarity)")
