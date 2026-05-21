import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def parse_log(text):
    """解析日志文本，提取统计信息和编号列表"""
    lines = [line.strip() for line in text.strip().split('\n') if line.strip()]

    if not lines:
        return None

    # 第一行通常是状态
    status = lines[0] if lines else ''

    # 第二行是统计摘要
    summary = {}
    if len(lines) > 1:
        match = re.search(r'All:(\d+),success:(\d+),fail:(\d+)', lines[1])
        if match:
            summary = {
                'total': int(match.group(1)),
                'success': int(match.group(2)),
                'fail': int(match.group(3))
            }

    # 剩余行是编号列表（通常是失败项）
    codes = []
    for line in lines[2:]:
        if line and not line.startswith('#'):
            codes.append(line)

    return {
        'status': status,
        'summary': summary,
        'codes': codes
    }


def create_excel(data, output_file='result.xlsx'):
    """生成Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "执行结果"

    # 样式定义
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    success_font = Font(color="006100")

    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "脚本执行结果统计"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 状态行
    ws['A2'] = "执行状态"
    ws['B2'] = data['status']
    ws['A2'].font = Font(bold=True)
    if data['status'].lower() == 'success':
        ws['B2'].fill = success_fill
        ws['B2'].font = success_font
    else:
        ws['B2'].fill = fail_fill
        ws['B2'].font = fail_font

    # 汇总统计
    ws['A4'] = "汇总统计"
    ws['A4'].font = Font(bold=True, size=12)

    headers = ["总任务数", "成功数", "失败数", "成功率"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    summary = data['summary']
    if summary:
        total = summary.get('total', 0)
        success = summary.get('success', 0)
        fail = summary.get('fail', 0)
        rate = f"{success/total*100:.1f}%" if total > 0 else "0%"

        ws.cell(row=6, column=1, value=total).border = thin_border
        ws.cell(row=6, column=2, value=success).border = thin_border
        success_cell = ws.cell(row=6, column=3, value=fail)
        success_cell.border = thin_border
        if fail > 0:
            success_cell.fill = fail_fill
            success_cell.font = fail_font
        ws.cell(row=6, column=4, value=rate).border = thin_border

    # 失败详情列表
    start_row = 8
    ws.cell(row=start_row, column=1, value="失败/异常编号列表")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    if data['codes']:
        ws.cell(row=start_row + 1, column=1, value="序号").font = header_font
        ws.cell(row=start_row + 1, column=1).fill = header_fill
        ws.cell(row=start_row + 1, column=1).alignment = header_align
        ws.cell(row=start_row + 1, column=1).border = thin_border

        ws.cell(row=start_row + 1, column=2, value="编号").font = header_font
        ws.cell(row=start_row + 1, column=2).fill = header_fill
        ws.cell(row=start_row + 1, column=2).alignment = header_align
        ws.cell(row=start_row + 1, column=2).border = thin_border

        for i, code in enumerate(data['codes'], 1):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=code).border = thin_border
    else:
        ws.cell(row=start_row + 1, column=1, value="无失败记录").font = Font(italic=True, color="999999")

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(output_file)
    print(f"Excel 已生成: {output_file}")
    print(f"  状态: {data['status']}")
    if data['summary']:
        print(f"  总计: {data['summary'].get('total', 0)}, "
              f"成功: {data['summary'].get('success', 0)}, "
              f"失败: {data['summary'].get('fail', 0)}")
    print(f"  失败编号数: {len(data['codes'])}")


def main():
    # 如果没有参数，使用示例数据
    if len(sys.argv) < 2:
        sample = """success
################All:19,success:12,fail:7#################
BZCHJC20260150
BZCHJC20240267
BZCHJC20260082
BZCHJC20260143
BZCHJC20260047
BZCHJC20260046
BZCHJC20260130"""
        print("未指定输入文件，使用示例数据...")
        print("用法: python convert_log_to_excel.py <日志文件.txt>")
        print()
        data = parse_log(sample)
        create_excel(data, 'result.xlsx')
        return

    input_file = sys.argv[1]
    output_file = input_file.rsplit('.', 1)[0] + '.xlsx'

    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            text = f.read()
    except Exception as e:
        print(f"读取文件失败: {e}")
        return

    data = parse_log(text)
    if data:
        create_excel(data, output_file)
    else:
        print("未能解析日志内容")


if __name__ == "__main__":
    main()
