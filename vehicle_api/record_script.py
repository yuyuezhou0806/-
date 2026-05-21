#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Selenium 操作录制工具
功能：记录你的鼠标点击、输入等操作，自动生成Python脚本
使用方法：
1. 运行这个脚本
2. 在打开的浏览器中操作你想自动化的流程
3. 操作完成后，在命令行按 Ctrl+C 结束录制
4. 生成的脚本会保存为 recorded_script.py
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
import json
import os
from datetime import datetime


class OperationRecorder:
    def __init__(self):
        self.operations = []
        self.driver = None
        self.start_time = None

    def start_recording(self, url='http://192.168.99.91/'):
        """开始录制"""
        print("=" * 70)
        print("🎥 Selenium 操作录制工具")
        print("=" * 70)
        print(f"\n1. 正在打开浏览器...")

        self.driver = webdriver.Chrome()
        self.driver.implicitly_wait(10)
        self.driver.get(url)

        print(f"2. 已打开: {url}")
        print(f"3. 请在浏览器中进行操作，系统会自动记录")
        print(f"4. 操作完成后，在本窗口按 Ctrl+C 结束录制\n")

        self.start_time = time.time()

        # 注入JavaScript监听用户操作
        self._inject_recorder()

        try:
            while True:
                time.sleep(1)
                # 每秒获取一次操作记录
                self._collect_operations()
        except KeyboardInterrupt:
            print("\n\n⏹️ 录制结束，正在生成脚本...")
        except Exception as e:
            print(f"\n\n⚠️ 发生异常: {e}")
            print("尝试生成已录制的脚本...")
        finally:
            # 最后尝试生成脚本（即使没有捕获到操作）
            self._generate_script()

    def _inject_recorder(self):
        """注入JavaScript代码监听用户操作"""
        script = """
        window.recordedOperations = [];

        // 生成元素的选择器
        function getSelector(element) {
            // 1. 优先使用ID
            if (element.id) {
                return 'By.ID, "' + element.id + '"';
            }

            // 2. 使用name属性
            if (element.name) {
                return 'By.NAME, "' + element.name + '"';
            }

            // 3. 使用placeholder
            if (element.placeholder) {
                return 'By.CSS_SELECTOR, "[placeholder=\\'' + element.placeholder + '\\']"';
            }

            // 4. 使用文本内容（按钮、链接）
            var text = element.innerText || element.textContent;
            if (text && text.trim() && (element.tagName === 'BUTTON' || element.tagName === 'A' || element.tagName === 'SPAN')) {
                text = text.trim().substring(0, 20);  // 限制长度
                return 'By.XPATH, "//' + element.tagName.toLowerCase() + '[contains(text(), \\\'' + text + '\\\')]"';
            }

            // 5. 使用class
            if (element.className && typeof element.className === 'string') {
                var classes = element.className.split(' ').filter(c => c && !c.includes('anticon')).slice(0, 2);
                if (classes.length > 0) {
                    return 'By.CSS_SELECTOR, ".' + classes.join('.') + '"';
                }
            }

            // 6. 使用tag name
            return 'By.TAG_NAME, "' + element.tagName.toLowerCase() + '"';
        }

        // 获取元素描述
        function getElementInfo(element) {
            var tag = element.tagName;
            var text = (element.innerText || element.textContent || '').trim().substring(0, 30);
            var elemId = element.id || '';
            var placeholder = element.placeholder || '';

            return {
                tag: tag,
                text: text,
                id: elemId,
                placeholder: placeholder,
                selector: getSelector(element)
            };
        }

        // 监听点击事件
        document.addEventListener('click', function(e) {
            var info = getElementInfo(e.target);
            window.recordedOperations.push({
                type: 'click',
                timestamp: Date.now(),
                element: info
            });
            console.log('🖱️ 点击:', info.text || info.tag, info.selector);
        }, true);

        // 监听输入事件
        document.addEventListener('input', function(e) {
            if (e.target.tagName === 'INPUT' || e.target.tagName === 'TEXTAREA') {
                var info = getElementInfo(e.target);
                window.recordedOperations.push({
                    type: 'input',
                    timestamp: Date.now(),
                    element: info,
                    value: e.target.value
                });
                console.log('⌨️ 输入:', info.placeholder || info.tag, '→', e.target.value.substring(0, 20));
            }
        }, true);

        // 监听双击事件
        document.addEventListener('dblclick', function(e) {
            var info = getElementInfo(e.target);
            window.recordedOperations.push({
                type: 'double_click',
                timestamp: Date.now(),
                element: info
            });
            console.log('🖱️🖱️ 双击:', info.text || info.tag);
        }, true);

        console.log('✅ 录制器已启动');
        """
        self.driver.execute_script(script)

    def _collect_operations(self):
        """收集JavaScript记录的操作"""
        try:
            operations = self.driver.execute_script("return window.recordedOperations || [];")

            # 只处理新记录的操作
            for op in operations[len(self.operations):]:
                # 添加等待时间
                if self.operations:
                    last_time = self.operations[-1].get('timestamp', 0)
                    wait_time = (op['timestamp'] - last_time) / 1000
                    if wait_time > 0.5:  # 如果间隔超过0.5秒，记录等待
                        op['wait_before'] = min(wait_time, 3)  # 最多等3秒

                self.operations.append(op)

        except Exception as e:
            pass

    def _generate_script(self):
        """生成Python脚本"""
        # 最后一次收集操作
        try:
            self._collect_operations()
        except:
            pass

        if not self.operations:
            print("⚠️ 没有记录到任何操作")
            print("可能原因：")
            print("  1. 浏览器中没有任何点击或输入操作")
            print("  2. JavaScript注入失败")
            print("\n建议：重新运行脚本，在浏览器中多点击几次元素")
            return

        script_lines = [
            "#!/usr/bin/env python3",
            "# -*- coding: utf-8 -*-",
            f"# 自动生成的脚本 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "# 针对卡顿网站的优化版本 - 包含多处 time.sleep 等待",
            "",
            "from selenium import webdriver",
            "from selenium.webdriver.common.by import By",
            "from selenium.webdriver.common.action_chains import ActionChains",
            "from selenium.webdriver.support.ui import WebDriverWait",
            "from selenium.webdriver.support import expected_conditions as EC",
            "import time",
            "",
            "# ========== 延时配置（根据网站响应速度调整） ==========",
            "CLICK_WAIT = 1.5      # 点击后等待（秒）",
            "INPUT_WAIT = 0.5      # 输入后等待（秒）",
            "PAGE_WAIT = 3.0       # 页面切换等待（秒）",
            "POPUP_WAIT = 2.0      # 弹窗等待（秒）",
            "MAX_RETRY = 3         # 失败重试次数",
            "# ====================================================",
            "",
            "",
            "def auto_operate(project_num='ZCHJC003869'):",
            '    """自动执行录制好的操作"""',
            '    """\n    参数:',
            '    project_num - 项目编号，默认为录制时使用的编号',
            '    如需处理多个项目，可传入不同的项目编号',
            '    """',
            "",
            "    driver = webdriver.Chrome()",
            "    driver.implicitly_wait(10)",
            "",
            f'    driver.get("{self.driver.current_url}")',
            "    wait = WebDriverWait(driver, 15)  # 最长等15秒",
            "",
            "    print('开始执行自动化操作...')",
            "",
        ]

        for i, op in enumerate(self.operations):
            # 添加等待（强制等待+智能等待）
            if 'wait_before' in op:
                wait_time = max(op['wait_before'], 1.0)  # 最少等1秒
                script_lines.append(f"    time.sleep(max({wait_time:.1f}, PAGE_WAIT))  # 等待页面响应")

            elem = op['element']
            selector = elem.get('selector', '')

            if op['type'] == 'click':
                script_lines.append(f"")
                script_lines.append(f"    # 点击: {elem.get('text', elem.get('tag', '元素'))}")
                script_lines.append(f"    wait.until(EC.element_to_be_clickable(({selector}))).click()")
                script_lines.append(f"    time.sleep(CLICK_WAIT)  # 点击后等待页面加载")

            elif op['type'] == 'input':
                value = op.get('value', '')
                script_lines.append(f"")
                script_lines.append(f"    # 输入: {elem.get('placeholder', elem.get('tag', '输入框'))}")
                script_lines.append(f"    input_elem = wait.until(EC.presence_of_element_located(({selector})))")
                script_lines.append(f"    input_elem.clear()")
                script_lines.append(f'    input_elem.send_keys("{value}")')
                script_lines.append(f"    time.sleep(INPUT_WAIT)  # 输入后等待")

            elif op['type'] == 'double_click':
                script_lines.append(f"")
                script_lines.append(f"    # 双击: {elem.get('text', elem.get('tag', '元素'))}")
                script_lines.append(f"    elem = wait.until(EC.presence_of_element_located(({selector})))")
                script_lines.append(f"    ActionChains(driver).double_click(elem).perform()")
                script_lines.append(f"    time.sleep(POPUP_WAIT)  # 双击后等待弹窗")

        # 添加结尾
        script_lines.extend([
            "",
            "",
            "    print('✅ 操作完成')",
            "    time.sleep(5)  # 完成后停留5秒查看结果",
            "    driver.quit()",
            "",
            "",
            "def batch_process(excel_file):",
            '    """批量处理Excel中的项目"""',
            '    """',
            '    示例Excel格式：',
            '    第一列：项目编号',
            '    第二列：工作量（可选）',
            '    """',
            "    import openpyxl",
            "    wb = openpyxl.load_workbook(excel_file)",
            "    sheet = wb.active",
            "",
            "    for row in range(2, sheet.max_row + 1):",
            "        project_num = sheet.cell(row=row, column=1).value",
            "        if project_num:",
            "            print(f'正在处理: {project_num}')",
            "            try:",
            "                auto_operate(project_num)",
            "                print(f'✅ {project_num} 完成')",
            "            except Exception as e:",
            "                print(f'❌ {project_num} 失败: {e}')",
            "",
            "",
            'if __name__ == "__main__":',
            "    import sys",
            "",
            "    if len(sys.argv) > 1:",
            "        # 命令行传入项目编号",
            "        auto_operate(sys.argv[1])",
            "    else:",
            "        # 使用默认项目编号",
            "        auto_operate()",
        ])

        # 保存脚本
        script_content = '\n'.join(script_lines)

        # 保存为文件
        filename = f'recorded_script_{datetime.now().strftime("%H%M%S")}.py'
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(script_content)

        # 同时保存原始记录
        with open(f'operations_{datetime.now().strftime("%H%M%S")}.json', 'w', encoding='utf-8') as f:
            json.dump(self.operations, f, ensure_ascii=False, indent=2)

        print(f"\n{'='*70}")
        print(f"✅ 脚本已生成: {filename}")
        print(f"✅ 原始记录已保存: operations_*.json")
        print(f"\n共记录 {len(self.operations)} 个操作:")

        # 统计操作类型
        click_count = sum(1 for op in self.operations if op['type'] == 'click')
        input_count = sum(1 for op in self.operations if op['type'] == 'input')
        double_click_count = sum(1 for op in self.operations if op['type'] == 'double_click')

        print(f"  - 点击: {click_count} 次")
        print(f"  - 输入: {input_count} 次")
        print(f"  - 双击: {double_click_count} 次")
        print(f"\n使用方法:")
        print(f"  python {filename}")
        print(f"{'='*70}\n")

        # 显示脚本预览
        print("📄 脚本预览 (前30行):")
        print("-"*70)
        print('\n'.join(script_lines[:30]))
        print("\n...")

    def close(self):
        """关闭浏览器"""
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass  # 浏览器可能已经关闭，忽略错误


if __name__ == '__main__':
    print("\n" + "="*70)
    print("🎥 Selenium 操作录制工具")
    print("="*70)
    print("\n使用步骤:")
    print("1. 本工具会自动打开 Chrome 浏览器并访问目标网页")
    print("2. 你在浏览器中的点击、输入等操作会被自动记录")
    print("3. 操作完成后，在本窗口按 Ctrl+C 结束录制")
    print("4. 自动生成 Python 脚本文件\n")

    url = input("请输入要录制的网页地址 (直接回车使用默认 http://192.168.99.91/): ").strip()
    if not url:
        url = 'http://192.168.99.91/'

    recorder = OperationRecorder()
    try:
        recorder.start_recording(url)
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
    finally:
        recorder.close()
        input("\n按回车键退出...")
