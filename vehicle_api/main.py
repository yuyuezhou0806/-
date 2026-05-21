#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
车辆档案管理系统 - 后端API
技术栈: FastAPI + SQLAlchemy + SQLite
"""

from datetime import datetime, date
from typing import List, Optional
from enum import Enum
from io import BytesIO

from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Date, Enum as SQLEnum
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io


# ==================== 数据库配置 ====================
SQLALCHEMY_DATABASE_URL = "sqlite:///./vehicle_management.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


# ==================== 数据库模型 ====================
class VehicleStatus(str, Enum):
    """车辆状态"""
    IDLE = "idle"           # 闲置
    IN_USE = "in_use"       # 使用中
    MAINTENANCE = "maintenance"  # 维修中
    SCRAPPED = "scrapped"   # 已报废


class Vehicle(Base):
    """车辆表"""
    __tablename__ = "vehicles"

    id = Column(Integer, primary_key=True, index=True)
    plate_number = Column(String(20), unique=True, nullable=False, comment="车牌号")
    brand = Column(String(50), comment="品牌")
    model = Column(String(50), comment="型号")
    color = Column(String(20), comment="颜色")
    seat_count = Column(Integer, default=5, comment="座位数")

    # 购置信息
    purchase_date = Column(Date, comment="购买日期")
    purchase_price = Column(Float, comment="购买价格")
    supplier = Column(String(100), comment="供应商")
    warranty_period = Column(Integer, comment="质保期(月)")

    # 证件信息
    vin_code = Column(String(50), comment="车架号")
    engine_number = Column(String(50), comment="发动机号")

    # 保险信息
    insurance_company = Column(String(50), comment="保险公司")
    insurance_no = Column(String(50), comment="保单号")
    insurance_start = Column(Date, comment="保险生效日期")
    insurance_end = Column(Date, comment="保险到期日期")

    # 状态
    status = Column(String(20), default="idle", comment="车辆状态")
    current_mileage = Column(Integer, default=0, comment="当前里程")
    department = Column(String(50), comment="所属部门")

    # 保养信息
    maintenance_interval = Column(Integer, default=5000, comment="保养周期(公里)")
    last_maintenance_mileage = Column(Integer, default=0, comment="上次保养里程")
    last_maintenance_date = Column(Date, comment="上次保养日期")

    # 时间戳
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)


class MaintenanceRecord(Base):
    """维修记录表"""
    __tablename__ = "maintenance_records"

    id = Column(Integer, primary_key=True, index=True)
    vehicle_id = Column(Integer, nullable=False, comment="车辆ID")
    maintenance_type = Column(String(50), comment="维修类型")
    description = Column(String(500), comment="维修描述")
    cost = Column(Float, default=0, comment="维修费用")
    maintenance_date = Column(Date, comment="维修日期")
    next_maintenance_date = Column(Date, comment="下次保养日期")
    mileage = Column(Integer, comment="维修时里程")
    repair_shop = Column(String(100), comment="维修店/4S店")
    status = Column(String(20), default="completed", comment="状态")
    created_at = Column(DateTime, default=datetime.now)


class VehicleApplication(Base):
    """用车申请表"""
    __tablename__ = "vehicle_applications"

    id = Column(Integer, primary_key=True, index=True)
    applicant_name = Column(String(50), nullable=False, comment="申请人")
    department = Column(String(50), comment="部门")
    phone = Column(String(20), comment="联系电话")
    reason = Column(String(500), comment="用车事由")
    destination = Column(String(200), comment="目的地")
    passengers = Column(Integer, default=1, comment="乘车人数")
    start_time = Column(DateTime, comment="预计出发时间")
    end_time = Column(DateTime, comment="预计返回时间")
    vehicle_id = Column(Integer, comment="分配的车辆ID")
    status = Column(String(20), default="pending", comment="状态: pending/approved/rejected/completed")
    approver = Column(String(50), comment="审批人")
    approval_time = Column(DateTime, comment="审批时间")
    approval_comment = Column(String(500), comment="审批意见")
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)


class User(Base):
    """用户表"""
    __tablename__ = "users"

    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, nullable=False, comment="用户名")
    password_hash = Column(String(128), nullable=False, comment="密码哈希")
    name = Column(String(50), comment="姓名")
    role = Column(String(20), default="user", comment="角色: admin/user")
    is_active = Column(Integer, default=1, comment="是否启用: 1/0")
    created_at = Column(DateTime, default=datetime.now)
    updated_at = Column(DateTime, default=datetime.now, onupdate=datetime.now)


class Token(Base):
    """登录Token表"""
    __tablename__ = "tokens"

    id = Column(Integer, primary_key=True, index=True)
    token = Column(String(128), unique=True, nullable=False, comment="Token字符串")
    user_id = Column(Integer, nullable=False, comment="用户ID")
    username = Column(String(50), comment="用户名")
    role = Column(String(20), comment="角色")
    expires_at = Column(DateTime, nullable=False, comment="过期时间")
    created_at = Column(DateTime, default=datetime.now, comment="创建时间")


# 创建表
Base.metadata.create_all(bind=engine)


# ==================== Pydantic 模型 ====================
class VehicleBase(BaseModel):
    """车辆基础模型"""
    plate_number: str = Field(..., min_length=5, max_length=20, description="车牌号")
    brand: Optional[str] = Field(None, max_length=50, description="品牌")
    model: Optional[str] = Field(None, max_length=50, description="型号")
    color: Optional[str] = Field(None, max_length=20, description="颜色")
    seat_count: Optional[int] = Field(default=5, ge=1, le=50, description="座位数")

    # 购置信息
    purchase_date: Optional[date] = Field(None, description="购买日期")
    purchase_price: Optional[float] = Field(None, ge=0, description="购买价格")
    supplier: Optional[str] = Field(None, max_length=100, description="供应商")
    warranty_period: Optional[int] = Field(None, ge=0, description="质保期(月)")

    # 证件信息
    vin_code: Optional[str] = Field(None, max_length=50, description="车架号")
    engine_number: Optional[str] = Field(None, max_length=50, description="发动机号")

    # 保险信息
    insurance_company: Optional[str] = Field(None, max_length=50, description="保险公司")
    insurance_no: Optional[str] = Field(None, max_length=50, description="保单号")
    insurance_start: Optional[date] = Field(None, description="保险生效日期")
    insurance_end: Optional[date] = Field(None, description="保险到期日期")

    # 状态
    status: str = Field(default="idle", description="车辆状态: idle/in_use/maintenance/scrapped")
    current_mileage: int = Field(default=0, ge=0, description="当前里程")
    department: Optional[str] = Field(None, max_length=50, description="所属部门")

    # 保养信息
    maintenance_interval: Optional[int] = Field(default=5000, ge=0, description="保养周期(公里)")
    last_maintenance_mileage: Optional[int] = Field(default=0, ge=0, description="上次保养里程")
    last_maintenance_date: Optional[date] = Field(None, description="上次保养日期")


class VehicleCreate(VehicleBase):
    """创建车辆"""
    pass


class VehicleUpdate(BaseModel):
    """更新车辆"""
    plate_number: Optional[str] = Field(None, min_length=5, max_length=20, description="车牌号")
    brand: Optional[str] = None
    model: Optional[str] = None
    color: Optional[str] = None
    seat_count: Optional[int] = None
    purchase_date: Optional[date] = None
    purchase_price: Optional[float] = None
    supplier: Optional[str] = None
    warranty_period: Optional[int] = None
    vin_code: Optional[str] = None
    engine_number: Optional[str] = None
    insurance_company: Optional[str] = None
    insurance_no: Optional[str] = None
    insurance_start: Optional[date] = None
    insurance_end: Optional[date] = None
    status: Optional[str] = None
    current_mileage: Optional[int] = None
    department: Optional[str] = None

    # 保养信息
    maintenance_interval: Optional[int] = None
    last_maintenance_mileage: Optional[int] = None
    last_maintenance_date: Optional[date] = None


class VehicleResponse(VehicleBase):
    """车辆响应模型"""
    id: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class VehicleList(BaseModel):
    """车辆列表响应"""
    total: int
    items: List[VehicleResponse]


# 登录相关模型
class LoginRequest(BaseModel):
    """登录请求"""
    username: str = Field(..., min_length=3, max_length=50, description="用户名")
    password: str = Field(..., min_length=6, max_length=50, description="密码")


class LoginResponse(BaseModel):
    """登录响应"""
    success: bool
    token: Optional[str] = None
    message: str
    user: Optional[dict] = None


class UserCreate(BaseModel):
    """创建用户"""
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6, max_length=50)
    name: Optional[str] = Field(None, max_length=50)
    role: str = Field(default="user", pattern="^(admin|user)$")


class UserResponse(BaseModel):
    """用户响应"""
    id: int
    username: str
    name: Optional[str]
    role: str
    is_active: bool
    created_at: Optional[datetime]

    class Config:
        from_attributes = True


# ==================== FastAPI 应用 ====================
app = FastAPI(
    title="车辆管理系统API",
    description="车辆档案管理、用车申请、维修记录等",
    version="1.0.0"
)

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载静态文件
app.mount("/static", StaticFiles(directory="static"), name="static")


# 依赖注入：获取数据库会话
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# 初始化默认管理员
def init_default_user():
    """初始化默认管理员账号"""
    db = SessionLocal()
    try:
        # 检查是否已有用户
        user = db.query(User).filter(User.username == "admin").first()
        if not user:
            # 创建默认管理员账号: admin/admin123
            import hashlib
            password_hash = hashlib.sha256("admin123".encode()).hexdigest()
            admin = User(
                username="admin",
                password_hash=password_hash,
                name="管理员",
                role="admin",
                is_active=1
            )
            db.add(admin)
            db.commit()
            print("✓ 默认管理员账号已创建: admin / admin123")
    except Exception as e:
        print(f"初始化用户失败: {e}")
    finally:
        db.close()


# 清理过期 token
def cleanup_expired_tokens():
    """清理所有过期的 token"""
    db = SessionLocal()
    try:
        expired_tokens = db.query(Token).filter(Token.expires_at < datetime.now()).all()
        for token in expired_tokens:
            db.delete(token)
        db.commit()
        if expired_tokens:
            print(f"✓ 已清理 {len(expired_tokens)} 个过期 token")
    except Exception as e:
        print(f"清理过期 token 失败: {e}")
        db.rollback()
    finally:
        db.close()


# 启动时初始化
init_default_user()
cleanup_expired_tokens()


# ==================== 登录认证相关 ====================

from datetime import timedelta


def verify_token(token: str = Query(None), db: Session = Depends(get_db)):
    """验证 token"""
    if not token:
        raise HTTPException(status_code=401, detail="未登录或登录已过期")

    # 从数据库查询 token
    token_record = db.query(Token).filter(Token.token == token).first()
    if not token_record:
        raise HTTPException(status_code=401, detail="未登录或登录已过期")

    # 检查是否过期
    if datetime.now() > token_record.expires_at:
        # 删除过期 token
        db.delete(token_record)
        db.commit()
        raise HTTPException(status_code=401, detail="登录已过期，请重新登录")

    return {"user_id": token_record.user_id, "username": token_record.username, "role": token_record.role}


@app.post("/api/login", response_model=LoginResponse)
def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    """用户登录"""
    import hashlib
    import secrets

    # 查找用户
    user = db.query(User).filter(User.username == login_data.username).first()

    if not user:
        return LoginResponse(success=False, message="用户名或密码错误")

    if not user.is_active:
        return LoginResponse(success=False, message="账号已被禁用")

    # 验证密码
    password_hash = hashlib.sha256(login_data.password.encode()).hexdigest()
    if password_hash != user.password_hash:
        return LoginResponse(success=False, message="用户名或密码错误")

    # 生成 token
    token_str = secrets.token_urlsafe(32)

    # 设置过期时间（7天）
    expires_at = datetime.now() + timedelta(days=7)

    # 保存到数据库
    new_token = Token(
        token=token_str,
        user_id=user.id,
        username=user.username,
        role=user.role,
        expires_at=expires_at
    )
    db.add(new_token)
    db.commit()

    return LoginResponse(
        success=True,
        token=token_str,
        message="登录成功",
        user={
            "id": user.id,
            "username": user.username,
            "name": user.name,
            "role": user.role
        }
    )


@app.post("/api/logout")
def logout(token: str = Query(...), db: Session = Depends(get_db)):
    """用户登出"""
    # 从数据库删除 token
    token_record = db.query(Token).filter(Token.token == token).first()
    if token_record:
        db.delete(token_record)
        db.commit()
    return {"success": True, "message": "已登出"}


@app.get("/api/users", response_model=List[UserResponse])
def list_users(db: Session = Depends(get_db), current_user: dict = Depends(verify_token)):
    """获取用户列表（仅管理员）"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="无权限")

    users = db.query(User).all()
    return users


@app.post("/api/users", response_model=UserResponse)
def create_user(user_data: UserCreate, db: Session = Depends(get_db), current_user: dict = Depends(verify_token)):
    """创建用户（仅管理员）"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="无权限")

    import hashlib

    # 检查用户名是否已存在
    existing = db.query(User).filter(User.username == user_data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名已存在")

    # 创建用户
    password_hash = hashlib.sha256(user_data.password.encode()).hexdigest()
    new_user = User(
        username=user_data.username,
        password_hash=password_hash,
        name=user_data.name,
        role=user_data.role,
        is_active=1
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return new_user


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: dict = Depends(verify_token)):
    """删除用户（仅管理员，不能删除自己）"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="无权限")

    # 不能删除自己
    if current_user["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="不能删除当前登录账号")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 不能删除admin账号
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="不能删除系统管理员账号")

    db.delete(user)
    db.commit()

    return {"success": True, "message": "删除成功"}


# ==================== API 路由 ====================

@app.get("/", response_class=HTMLResponse)
def root():
    """根路径 - 返回前端页面"""
    with open("index.html", "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(content=content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/index.html", response_class=HTMLResponse)
def index():
    """前端页面"""
    with open("index.html", "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(content=content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/test.html", response_class=HTMLResponse)
def test():
    """测试页面"""
    with open("test.html", "r", encoding="utf-8") as f:
        content = f.read()
    return HTMLResponse(content=content, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


# ----- 车辆档案管理 -----

@app.post("/api/vehicles", response_model=VehicleResponse)
def create_vehicle(vehicle: VehicleCreate, db: Session = Depends(get_db)):
    """创建车辆档案"""
    # 检查车牌号是否已存在
    existing = db.query(Vehicle).filter(Vehicle.plate_number == vehicle.plate_number).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"车牌号 {vehicle.plate_number} 已存在")

    db_vehicle = Vehicle(**vehicle.dict())
    db.add(db_vehicle)
    db.commit()
    db.refresh(db_vehicle)
    return db_vehicle


@app.get("/api/vehicles", response_model=VehicleList)
def list_vehicles(
    skip: int = Query(0, ge=0, description="跳过数量"),
    limit: int = Query(20, ge=1, le=100, description="每页数量"),
    status: Optional[str] = Query(None, description="状态筛选"),
    department: Optional[str] = Query(None, description="部门筛选"),
    keyword: Optional[str] = Query(None, description="关键词搜索(车牌/品牌/型号)"),
    db: Session = Depends(get_db)
):
    """获取车辆列表"""
    query = db.query(Vehicle)

    # 筛选条件
    if status:
        query = query.filter(Vehicle.status == status)
    if department:
        query = query.filter(Vehicle.department == department)
    if keyword:
        query = query.filter(
            (Vehicle.plate_number.contains(keyword)) |
            (Vehicle.brand.contains(keyword)) |
            (Vehicle.model.contains(keyword))
        )

    total = query.count()
    items = query.order_by(Vehicle.created_at.desc()).offset(skip).limit(limit).all()

    return {"total": total, "items": items}


@app.get("/api/vehicles/export")
def export_vehicles(
    status: Optional[str] = Query(None, description="状态筛选"),
    department: Optional[str] = Query(None, description="部门筛选"),
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    db: Session = Depends(get_db)
):
    """导出车辆数据到Excel"""
    # 查询数据（不分页，导出全部）
    query = db.query(Vehicle)

    if status:
        query = query.filter(Vehicle.status == status)
    if department:
        query = query.filter(Vehicle.department == department)
    if keyword:
        query = query.filter(
            (Vehicle.plate_number.contains(keyword)) |
            (Vehicle.brand.contains(keyword)) |
            (Vehicle.model.contains(keyword))
        )

    vehicles = query.order_by(Vehicle.created_at.desc()).all()

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "车辆档案"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "序号", "车牌号", "品牌", "型号", "颜色", "座位数", "状态",
        "所属部门", "当前里程(km)", "购买日期", "购买价格(元)", "供应商",
        "保险公司", "保单号", "保险生效", "保险到期",
        "保养周期(km)", "上次保养里程", "上次保养日期",
        "车架号", "发动机号", "创建时间"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 状态映射
    status_map = {
        "idle": "闲置",
        "in_use": "使用中",
        "maintenance": "维修中",
        "scrapped": "已报废"
    }

    # 填充数据
    for row_idx, vehicle in enumerate(vehicles, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=vehicle.plate_number)
        ws.cell(row=row_idx, column=3, value=vehicle.brand)
        ws.cell(row=row_idx, column=4, value=vehicle.model)
        ws.cell(row=row_idx, column=5, value=vehicle.color)
        ws.cell(row=row_idx, column=6, value=vehicle.seat_count)
        ws.cell(row=row_idx, column=7, value=status_map.get(vehicle.status, vehicle.status))
        ws.cell(row=row_idx, column=8, value=vehicle.department)
        ws.cell(row=row_idx, column=9, value=vehicle.current_mileage)
        ws.cell(row=row_idx, column=10, value=vehicle.purchase_date.strftime("%Y-%m-%d") if vehicle.purchase_date else "")
        ws.cell(row=row_idx, column=11, value=vehicle.purchase_price)
        ws.cell(row=row_idx, column=12, value=vehicle.supplier)
        ws.cell(row=row_idx, column=13, value=vehicle.insurance_company)
        ws.cell(row=row_idx, column=14, value=vehicle.insurance_no)
        ws.cell(row=row_idx, column=15, value=vehicle.insurance_start.strftime("%Y-%m-%d") if vehicle.insurance_start else "")
        ws.cell(row=row_idx, column=16, value=vehicle.insurance_end.strftime("%Y-%m-%d") if vehicle.insurance_end else "")
        ws.cell(row=row_idx, column=17, value=vehicle.maintenance_interval)
        ws.cell(row=row_idx, column=18, value=vehicle.last_maintenance_mileage)
        ws.cell(row=row_idx, column=19, value=vehicle.last_maintenance_date.strftime("%Y-%m-%d") if vehicle.last_maintenance_date else "")
        ws.cell(row=row_idx, column=20, value=vehicle.vin_code)
        ws.cell(row=row_idx, column=21, value=vehicle.engine_number)
        ws.cell(row=row_idx, column=22, value=vehicle.created_at.strftime("%Y-%m-%d %H:%M:%S") if vehicle.created_at else "")

        # 添加边框
        for col in range(1, 23):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 调整列宽
    column_widths = [6, 15, 12, 15, 10, 8, 10, 15, 12, 12, 12, 20, 15, 20, 12, 12, 12, 12, 12, 25, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名（使用URL编码支持中文）
    from urllib.parse import quote
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"车辆档案_{timestamp}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.post("/api/vehicles/import")
def import_vehicles(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """从Excel导入车辆数据"""
    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 .xlsx 或 .xls 格式的文件")

    try:
        # 读取上传的文件内容
        content = file.file.read()
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active

        # 验证表头
        expected_headers = ["车牌号", "品牌", "型号", "颜色", "座位数", "状态", "所属部门"]
        actual_headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]

        if not all(h in actual_headers for h in expected_headers[:3]):
            raise HTTPException(status_code=400, detail="文件格式不正确，请使用导出的模板文件")

        # 状态映射（中文 -> 英文）
        status_map = {
            "闲置": "idle",
            "使用中": "in_use",
            "维修中": "maintenance",
            "已报废": "scrapped"
        }

        imported_count = 0
        skipped_count = 0
        errors = []

        # 从第2行开始读取数据
        for row_idx in range(2, ws.max_row + 1):
            try:
                # 读取车牌号（必填）
                plate_number = ws.cell(row=row_idx, column=2).value  # B列
                if not plate_number:
                    continue

                plate_number = str(plate_number).strip()

                # 检查车牌号是否已存在
                existing = db.query(Vehicle).filter(Vehicle.plate_number == plate_number).first()
                if existing:
                    skipped_count += 1
                    continue

                # 读取其他字段
                brand = str(ws.cell(row=row_idx, column=3).value or '')
                model = str(ws.cell(row=row_idx, column=4).value or '')
                color = str(ws.cell(row=row_idx, column=5).value or '')

                # 座位数
                seat_count = ws.cell(row=row_idx, column=6).value
                try:
                    seat_count = int(seat_count) if seat_count else 5
                except:
                    seat_count = 5

                # 状态
                status_text = ws.cell(row=row_idx, column=7).value or '闲置'
                status = status_map.get(str(status_text).strip(), 'idle')

                # 所属部门
                department = str(ws.cell(row=row_idx, column=8).value or '')

                # 当前里程
                mileage = ws.cell(row=row_idx, column=9).value
                try:
                    current_mileage = int(float(mileage)) if mileage else 0
                except:
                    current_mileage = 0

                # 创建车辆记录
                vehicle = Vehicle(
                    plate_number=plate_number,
                    brand=brand,
                    model=model,
                    color=color,
                    seat_count=seat_count,
                    status=status,
                    department=department,
                    current_mileage=current_mileage
                )

                db.add(vehicle)
                imported_count += 1

            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")

        db.commit()

        return {
            "success": True,
            "imported": imported_count,
            "skipped": skipped_count,
            "errors": errors[:10]  # 最多返回10个错误
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@app.get("/api/vehicles/template")
def download_template():
    """下载导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "车辆导入模板"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "序号", "车牌号*", "品牌", "型号", "颜色", "座位数", "状态", "所属部门", "当前里程(km)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 示例数据
    examples = [
        [1, "京A12345", "奥迪", "A6L", "黑色", 5, "闲置", "行政部", 50000],
        [2, "沪AD12345", "特斯拉", "Model 3", "白色", 5, "使用中", "销售部", 30000],
    ]

    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 调整列宽
    column_widths = [8, 15, 12, 15, 10, 10, 10, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = "车辆导入模板.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/vehicles/{vehicle_id}", response_model=VehicleResponse)
def get_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    """获取单个车辆详情"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="车辆不存在")
    return vehicle


@app.put("/api/vehicles/{vehicle_id}", response_model=VehicleResponse)
def update_vehicle(vehicle_id: int, vehicle_update: VehicleUpdate, db: Session = Depends(get_db)):
    """更新车辆信息"""
    db_vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not db_vehicle:
        raise HTTPException(status_code=404, detail="车辆不存在")

    update_data = vehicle_update.dict(exclude_unset=True)

    # 如果修改了车牌号，检查是否已存在
    if 'plate_number' in update_data and update_data['plate_number']:
        new_plate = update_data['plate_number']
        if new_plate != db_vehicle.plate_number:
            existing = db.query(Vehicle).filter(Vehicle.plate_number == new_plate).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"车牌号 {new_plate} 已存在")

    for key, value in update_data.items():
        setattr(db_vehicle, key, value)

    db_vehicle.updated_at = datetime.now()
    db.commit()
    db.refresh(db_vehicle)
    return db_vehicle


@app.delete("/api/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    """删除车辆档案"""
    db_vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not db_vehicle:
        raise HTTPException(status_code=404, detail="车辆不存在")

    db.delete(db_vehicle)
    db.commit()
    return {"message": "删除成功"}


# ----- 统计信息 -----

@app.get("/api/statistics")
def get_statistics(db: Session = Depends(get_db)):
    """获取车辆统计信息"""
    total = db.query(Vehicle).count()
    idle = db.query(Vehicle).filter(Vehicle.status == "idle").count()
    in_use = db.query(Vehicle).filter(Vehicle.status == "in_use").count()
    maintenance = db.query(Vehicle).filter(Vehicle.status == "maintenance").count()
    scrapped = db.query(Vehicle).filter(Vehicle.status == "scrapped").count()

    # 保险即将到期（30天内）
    from datetime import timedelta
    today = date.today()
    soon_expire = db.query(Vehicle).filter(
        Vehicle.insurance_end <= today + timedelta(days=30),
        Vehicle.insurance_end >= today
    ).count()

    return {
        "total": total,
        "idle": idle,
        "in_use": in_use,
        "maintenance": maintenance,
        "scrapped": scrapped,
        "insurance_soon_expire": soon_expire
    }


# ----- 维修记录管理 -----

class MaintenanceRecordBase(BaseModel):
    """维修记录基础模型"""
    vehicle_id: int = Field(..., description="车辆ID")
    maintenance_type: str = Field(..., description="维修类型")
    description: Optional[str] = Field(None, description="维修描述")
    cost: float = Field(default=0, ge=0, description="维修费用")
    maintenance_date: Optional[date] = Field(None, description="维修日期")
    next_maintenance_date: Optional[date] = Field(None, description="下次保养日期")
    mileage: Optional[int] = Field(None, description="维修时里程")
    repair_shop: Optional[str] = Field(None, description="维修店/4S店")
    status: str = Field(default="completed", description="状态")

class MaintenanceRecordCreate(MaintenanceRecordBase):
    """创建维修记录"""
    pass

class MaintenanceRecordUpdate(BaseModel):
    """更新维修记录"""
    maintenance_type: Optional[str] = None
    description: Optional[str] = None
    cost: Optional[float] = None
    maintenance_date: Optional[date] = None
    next_maintenance_date: Optional[date] = None
    mileage: Optional[int] = None
    repair_shop: Optional[str] = None
    status: Optional[str] = None

class MaintenanceRecordResponse(MaintenanceRecordBase):
    """维修记录响应模型"""
    id: int
    created_at: datetime

    class Config:
        from_attributes = True

class MaintenanceList(BaseModel):
    """维修记录列表响应"""
    total: int
    items: List[MaintenanceRecordResponse]


# ----- 用车申请模型 -----

class VehicleApplicationBase(BaseModel):
    """用车申请基础模型"""
    applicant_name: str = Field(..., description="申请人")
    department: Optional[str] = Field(None, description="部门")
    phone: Optional[str] = Field(None, description="联系电话")
    reason: Optional[str] = Field(None, description="用车事由")
    destination: Optional[str] = Field(None, description="目的地")
    passengers: int = Field(default=1, ge=1, description="乘车人数")
    start_time: Optional[datetime] = Field(None, description="预计出发时间")
    end_time: Optional[datetime] = Field(None, description="预计返回时间")

class VehicleApplicationCreate(VehicleApplicationBase):
    """创建用车申请"""
    pass

class VehicleApplicationUpdate(BaseModel):
    """更新用车申请"""
    applicant_name: Optional[str] = None
    department: Optional[str] = None
    phone: Optional[str] = None
    reason: Optional[str] = None
    destination: Optional[str] = None
    passengers: Optional[int] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None

class VehicleApplicationResponse(VehicleApplicationBase):
    """用车申请响应模型"""
    id: int
    vehicle_id: Optional[int] = None
    status: str
    approver: Optional[str] = None
    approval_time: Optional[datetime] = None
    approval_comment: Optional[str] = None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class VehicleApplicationList(BaseModel):
    """用车申请列表响应"""
    total: int
    items: List[VehicleApplicationResponse]

class ApplicationApproval(BaseModel):
    """申请审批"""
    vehicle_id: Optional[int] = Field(None, description="分配的车辆ID")
    approver: str = Field(..., description="审批人")
    status: str = Field(..., description="approved/rejected")
    comment: Optional[str] = Field(None, description="审批意见")


# ----- 维修记录API -----

@app.post("/api/maintenance", response_model=MaintenanceRecordResponse)
def create_maintenance(record: MaintenanceRecordCreate, db: Session = Depends(get_db)):
    """创建维修记录"""
    vehicle = db.query(Vehicle).filter(Vehicle.id == record.vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="车辆不存在")

    db_record = MaintenanceRecord(**record.dict())
    db.add(db_record)
    db.commit()
    db.refresh(db_record)

    vehicle.status = "maintenance"
    vehicle.updated_at = datetime.now()
    db.commit()

    return db_record


@app.get("/api/maintenance", response_model=MaintenanceList)
def list_maintenance(
    vehicle_id: Optional[int] = Query(None, description="车辆ID筛选"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取维修记录列表"""
    query = db.query(MaintenanceRecord)
    if vehicle_id:
        query = query.filter(MaintenanceRecord.vehicle_id == vehicle_id)

    total = query.count()
    items = query.order_by(MaintenanceRecord.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": items}


@app.get("/api/maintenance/{record_id}", response_model=MaintenanceRecordResponse)
def get_maintenance(record_id: int, db: Session = Depends(get_db)):
    """获取单个维修记录"""
    record = db.query(MaintenanceRecord).filter(MaintenanceRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="维修记录不存在")
    return record


@app.delete("/api/maintenance/{record_id}")
def delete_maintenance(record_id: int, db: Session = Depends(get_db)):
    """删除维修记录"""
    db_record = db.query(MaintenanceRecord).filter(MaintenanceRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="维修记录不存在")

    db.delete(db_record)
    db.commit()
    return {"message": "删除成功"}


# ----- 用车申请API -----

@app.post("/api/applications", response_model=VehicleApplicationResponse)
def create_application(application: VehicleApplicationCreate, db: Session = Depends(get_db)):
    """创建用车申请"""
    db_app = VehicleApplication(**application.dict())
    db.add(db_app)
    db.commit()
    db.refresh(db_app)
    return db_app


@app.get("/api/applications", response_model=VehicleApplicationList)
def list_applications(
    status: Optional[str] = Query(None, description="状态筛选"),
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取用车申请列表"""
    query = db.query(VehicleApplication)
    if status:
        query = query.filter(VehicleApplication.status == status)

    total = query.count()
    items = query.order_by(VehicleApplication.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": items}


@app.get("/api/applications/{app_id}", response_model=VehicleApplicationResponse)
def get_application(app_id: int, db: Session = Depends(get_db)):
    """获取单个用车申请"""
    app = db.query(VehicleApplication).filter(VehicleApplication.id == app_id).first()
    if not app:
        raise HTTPException(status_code=404, detail="申请不存在")
    return app


@app.put("/api/applications/{app_id}/approve")
def approve_application(app_id: int, approval: ApplicationApproval, db: Session = Depends(get_db)):
    """审批用车申请"""
    app = db.query(VehicleApplication).filter(VehicleApplication.id == app_id).first()
    if not app:
        raise HTTPException(status_code=404, detail="申请不存在")

    if app.status != "pending":
        raise HTTPException(status_code=400, detail="该申请已处理")

    # 如果分配了车辆，检查车辆是否可用
    if approval.vehicle_id:
        vehicle = db.query(Vehicle).filter(Vehicle.id == approval.vehicle_id).first()
        if not vehicle:
            raise HTTPException(status_code=404, detail="车辆不存在")
        if vehicle.status != "idle":
            raise HTTPException(status_code=400, detail="该车辆当前不可用")

        # 更新车辆状态为使用中
        vehicle.status = "in_use"
        app.vehicle_id = approval.vehicle_id

    app.status = approval.status
    app.approver = approval.approver
    app.approval_time = datetime.now()
    app.approval_comment = approval.comment

    db.commit()
    db.refresh(app)
    return {"message": "审批成功", "application": VehicleApplicationResponse.from_orm(app)}


@app.delete("/api/applications/{app_id}")
def delete_application(app_id: int, db: Session = Depends(get_db)):
    """删除用车申请"""
    app = db.query(VehicleApplication).filter(VehicleApplication.id == app_id).first()
    if not app:
        raise HTTPException(status_code=404, detail="申请不存在")

    # 如果已分配车辆，释放车辆
    if app.vehicle_id and app.status == "approved":
        vehicle = db.query(Vehicle).filter(Vehicle.id == app.vehicle_id).first()
        if vehicle:
            vehicle.status = "idle"

    db.delete(app)
    db.commit()
    return {"message": "删除成功"}


@app.get("/api/vehicles/alerts/insurance")
def get_insurance_alerts(db: Session = Depends(get_db)):
    """获取保险即将到期的车辆列表"""
    from datetime import timedelta
    today = date.today()

    vehicles = db.query(Vehicle).filter(
        Vehicle.insurance_end <= today + timedelta(days=30),
        Vehicle.insurance_end >= today
    ).order_by(Vehicle.insurance_end).all()

    return {
        "count": len(vehicles),
        "items": [VehicleResponse.from_orm(v) for v in vehicles]
    }


# ----- 数据备份 -----

@app.get("/api/backup")
def backup_database():
    """备份数据库"""
    import shutil
    from datetime import datetime

    # 源数据库文件
    db_file = "./vehicle_management.db"
    backup_dir = "./backups"

    # 创建备份目录
    os.makedirs(backup_dir, exist_ok=True)

    # 生成备份文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(backup_dir, f"vehicle_management_backup_{timestamp}.db")

    # 复制数据库文件
    shutil.copy2(db_file, backup_file)

    return {
        "success": True,
        "message": f"备份成功",
        "file": backup_file,
        "timestamp": timestamp
    }


# ==================== 启动 ====================
if __name__ == "__main__":
    import uvicorn
    import os
    uvicorn.run(app, host="0.0.0.0", port=8000)
