from datetime import datetime
import openpyxl
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def main_point(project_num):
    """
    :param project_num:项目编号
    :return: 布尔型，返回运行结果
    """
    driver = None
    try:
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)
        driver.get('http://192.168.99.91/')

        # ===== 登录页面 =====
        # 原：/html/body/app-root/layout-passport/div/div/passport-login/form/nz-form-item[2]/button
        # 改：找placeholder包含"用户名"或"账号"的输入框
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder*="用户名"], input[placeholder*="账号"], .ant-input'))
        ).send_keys('Z0343')

        # 点击登录按钮（找包含"登录"文字的按钮）
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(., "登录") or contains(@class, "login")]'))
        ).click()

        # ===== 进入系统后 =====
        # 原：/html/body/app-root/layout-default/layout-header/div[2]/ul[1]/li[2]/header-app/nz-dropdown/div/span
        # 改：等待"自建应用"或"应用"字样出现并点击
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "自建应用") or contains(text(), "应用")]'))
        ).click()

        # 原：/html/body/div[3]/div[3]/div/div/div/div/nz-spin/div/div[1]/div[4]/i
        # 改：点击"项目管理"（基于文本）
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "项目管理")]'))
        ).click()

        # 原：/html/body/app-root/layout-default/section/app-inventory/div/div/define-project/div/as-split/as-split-area[2]/div[1]/div[1]/button[1]
        # 改：点击"查询方案"按钮
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(., "查询方案") or contains(@class, "search")]'))
        ).click()

        # 原：//*[@id="body"]/nz-layout/nz-content/nz-tabset/div[2]/div[1]/nz-table/nz-spin/div/div/div[1]/div/div/a[1]
        # 改：点击加号图标（添加过滤条件）
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.anticon-plus, [class*="plus"], button i.anticon'))
        ).click()

        # ===== 设置过滤条件 =====
        # 选择"项目编码"（第一个下拉框）
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '(//nz-select)[1]'))
        ).click()
        time.sleep(0.5)
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//li[contains(text(), "项目编码")]'))
        ).click()

        # 选择"等于"（第二个下拉框）
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '(//nz-select)[2]'))
        ).click()
        time.sleep(0.5)
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//li[contains(text(), "等于")]'))
        ).click()

        # 输入项目编号
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[placeholder*="值"], .ant-input'))
        ).send_keys(project_num)

        # 点击确定
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(., "确定") or contains(@class, "primary")]'))
        ).click()

        # ===== 选择项目 =====
        # 等待并点击查询结果中的项目编码
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, f'//td[contains(text(), "{project_num}")]'))
        ).click()

        # ===== 关联生成 =====
        # 点击"关联生成"下拉
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(., "关联生成") or contains(., "生成")]'))
        ).click()
        time.sleep(0.5)

        # 点击"生成项目工作汇报单"
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "项目工作汇报单") or contains(text(), "工作汇报")]'))
        ).click()

        # ===== 填写表单 =====
        # 等待弹窗出现
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.ant-modal, .pop-window, [class*="modal"]'))
        )

        # 选择某个下拉值（第18个表单项）
        # 使用通用方式：找包含"汇报"或"类型"的下拉框
        try:
            dropdown = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, '//nz-select[contains(., "汇报") or contains(., "类型") or position()=1]'))
            )
            dropdown.click()
            time.sleep(0.5)
            # 选择第3个选项
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, '(//li[contains(@class, "ant-select-dropdown-menu-item")])[3]'))
            ).click()
        except:
            pass  # 如果找不到就跳过

        # 点击新增按钮
        try:
            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(., "新增") or contains(@class, "add")]'))
            ).click()
        except:
            pass

        # ===== 填写明细 =====
        # 选择人员（双击）
        try:
            worker_cell = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//td[contains(@class, "ant-table-cell")][1]'))
            )
            ActionChains(driver).double_click(worker_cell).perform()
            time.sleep(0.5)
            # 选择第一个人
            first_worker = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//tbody/tr[1]/td[1]'))
            )
            ActionChains(driver).double_click(first_worker).perform()
        except:
            pass

        # 输入工时（0）
        try:
            hour_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input[nz-input-number], .ant-input-number-input'))
            )
            hour_input.clear()
            hour_input.send_keys('0')
        except:
            pass

        # 输入工作量（0）
        try:
            workload_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '(//input[@type="number"])[2]'))
            )
            workload_input.clear()
            workload_input.send_keys('0')
        except:
            pass

        # 点击审核/提交按钮
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(., "审核") or contains(., "提交") or contains(., "保存")]'))
        ).click()

        print('finish')
        time.sleep(2)
        return True

    except Exception as e:
        print(f'Error: {e}')
        # 出错时截图保存
        if driver:
            driver.save_screenshot(f'error_{project_num}_{datetime.now().strftime("%H%M%S")}.png')
        return False

    finally:
        if driver:
            driver.quit()


def read_excel(main_excel_name):
    a = []
    workbook = openpyxl.load_workbook(main_excel_name + '.xlsx')
    sheet = workbook.active
    for i in sheet.iter_rows():
        for j in i:
            a.append(j.value)
    return a


def write_log_data(log_file_name, log_data):
    with open(log_file_name, 'a', encoding='utf-8') as file:
        file.write(log_data)
        file.write('\n')


def yyz_main(main_excel_name):
    # 把表格名拼接出日志文件名
    log_file_name = 'log_' + main_excel_name + '.txt'
    # 建空列表记录运行错误的编号
    error_data = []
    # 读excel
    wb = openpyxl.load_workbook(main_excel_name + '.xlsx')
    # 默认sheet，如果指定sheet名就 sheet = wb.active('sheet名')
    sheet = wb.active
    # 因为第一行是标题，获取最大长度要-1，比如10行的excel就是9个数据,test_num=9
    test_num = sheet.max_row - 1

    print(f'共{test_num}条数据')

    # 从2开始遍历到9+2，因为python是左闭右开，所以代码是[2,11),从第二行读到第十行,一共9次
    for i in range(2, test_num + 2):
        # 获取对应行数和第1列的项目编号，和第2列的工作量
        project_num = sheet.cell(row=i, column=1).value

        # 打日志
        start_log = 'No:{},project_num:{},start_time:{}'.format(
            i - 1, project_num, time.strftime('%Y-%m-%d %H:%M:%S')
        )
        print(start_log)
        write_log_data(log_file_name, start_log)

        # 运行selenium脚本，把这次运行的project_num传进去
        flag = main_point(project_num)

        # 按返回类型打日志是否运行成功
        if flag:
            print('success')
            write_log_data(log_file_name, 'success')
        else:
            # 把运行错误的编号存到error_data列表
            error_data.append(project_num)
            print('fail!!')
            write_log_data(log_file_name, 'fail!!')

    # 那列表长度确定错误数
    fail_data_sum = len(error_data)
    fail_data_title = '####################All:{},success:{},fail:{}####################'.format(
        test_num, test_num - fail_data_sum, fail_data_sum
    )
    write_log_data(log_file_name, fail_data_title)
    for i in error_data:
        write_log_data(log_file_name, str(i))

    print(f'\n完成！成功:{test_num - fail_data_sum}, 失败:{fail_data_sum}')


if __name__ == '__main__':
    # 输入获取表格名，传到主函数
    excel_name = input('表格名（不带.xlsx）：')
    yyz_main(excel_name)
